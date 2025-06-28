from django.views.generic import TemplateView, View
from hr.models import *
from django.http import JsonResponse, HttpResponse
from io import BytesIO
from django.core.files.base import ContentFile
import qrcode
import json
from django.views import View
from User.models import *
from User.models import UserSession
import json
import uuid
from django.utils import timezone
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
import random
from datetime import datetime, timedelta
import base64
from django.shortcuts import redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import logging
from django.core.files.storage import default_storage
import os
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from hr.models import OnboardingInvitation
from django.contrib.auth.hashers import make_password
import string
from functools import wraps
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.db.models.signals import pre_save
import math
from xhtml2pdf import pisa
from django.template.loader import get_template, render_to_string
from django.core.mail import EmailMessage
import openpyxl
from openpyxl.utils import get_column_letter
from django.db import transaction

logger = logging.getLogger(__name__)

def class_employee_login_required(cls):
    original_dispatch = cls.dispatch
    
    def new_dispatch(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return redirect('hr_employee_login')
        return original_dispatch(self, request, *args, **kwargs)
    
    cls.dispatch = new_dispatch
    return cls

def get_user_from_session(request):
    """
    Get the user from the session.
    
    Args:
        request: The HTTP request object containing the session
        
    Returns:
        User object if authenticated, None otherwise
    """
    user_session_id = request.session.get('user_session_id')
    if not user_session_id:
        return None
    
    try:
        user_session = UserSession.objects.get(id=user_session_id)
        user = user_session.user
        user_id = user.user_id
        
        try:
            return User.objects.get(user_id=user_id)
        except User.DoesNotExist:
            logger.error(f"User with ID {user_id} does not exist")
            return None
            
    except UserSession.DoesNotExist:
        logger.warning("Invalid session in get_user_from_session")
        return None


# Create your views here.
class CompanyView(TemplateView):
    template_name = 'hr_management/company.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user = get_user_from_session(self.request)
        print(user)
        
        if user:
            context['companies'] = Company.objects.filter(user=user)
            context['qr_codes'] = QRCode.objects.filter(user=user)
        else:
            context['companies'] = Company.objects.all()
            context['qr_codes'] = QRCode.objects.all()
            
        return context

class CreationView(TemplateView):
    template_name = 'hr_management/creation.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user = None
        user_session_id = self.request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None
        
        if user:
            # Filter data by user
            context['departments'] = Department.objects.filter(user=user)
            context['designations'] = Designation.objects.filter(user=user)
            context['tandcs'] = TandC.objects.filter(user=user)
            context['roles'] = Role.objects.filter(user=user)
            
            # Get folders
            folders = Folder.objects.filter(user=user)
            context['folder_list'] = folders
            
            # Debug information
            print(f"User {user.user_id} has {folders.count()} folders")
            for folder in folders:
                print(f"Folder: {folder.folder_id}, {folder.name}")
            
            context['offer_letters'] = OfferLetter.objects.filter(user=user)
        else:
            # If no user is logged in, get all folders
            print("No user logged in, showing all folders")
            all_folders = Folder.objects.all()
            context['folder_list'] = all_folders
            print(f"Found {all_folders.count()} folders total")
            for folder in all_folders:
                print(f"Folder: {folder.folder_id}, {folder.name}")
                
            # Also get all other data
            context['departments'] = Department.objects.all()
            context['designations'] = Designation.objects.all()
            context['tandcs'] = TandC.objects.all()
            context['roles'] = Role.objects.all()
            context['offer_letters'] = OfferLetter.objects.all()
            
        return context

    def post(self, request, *args, **kwargs):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                from User.models import User
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            # Handle department creation
            if 'department_name' in request.POST:
                department = Department.objects.create(name=request.POST['department_name'])
                if user:
                    department.user = user
                    department.save()
                return JsonResponse({'success': True})

            # Handle designation creation
            elif 'designation_name' in request.POST:
                designation = Designation.objects.create(name=request.POST['designation_name'])
                if user:
                    designation.user = user
                    designation.save()
                return JsonResponse({'success': True})

            # Handle T&C creation
            elif 'tandc_name' in request.POST:
                tandc = TandC.objects.create(
                    name=request.POST['tandc_name'],
                    description=request.POST['tandc_description']
                )
                if user:
                    tandc.user = user
                    tandc.save()
                return JsonResponse({'success': True})

            # Handle role creation
            elif 'role_name' in request.POST:
                role = Role.objects.create(name=request.POST['role_name'])
                if user:
                    role.user = user
                    role.save()
                return JsonResponse({'success': True})

            # Handle offer letter creation
            elif 'offer_letter_name' in request.POST:
                offer_letter = OfferLetter.objects.create(
                    name=request.POST['offer_letter_name'],
                    content=request.POST['offer_letter_content']
                )
                return JsonResponse({'success': True})

            return JsonResponse({'error': 'Invalid form data'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class OnboardingView(TemplateView):
    template_name = 'hr_management/onboarding.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = get_user_from_session(self.request)
        
        user_companies = Company.objects.filter(user=user)
        context['companies'] = user_companies

        context['employees'] = Employee.objects.filter(user=user)
        context['offer_letters'] = OfferLetter.objects.filter(user=user)
        context['roles'] = Role.objects.filter(user=user)
        context['departments'] = Department.objects.filter(user=user)
        context['designations'] = Designation.objects.filter(user=user)
        context['tandcs'] = TandC.objects.filter(user=user)
        context['hiring_agreements'] = HiringAgreement.objects.filter(user=user)
        context['handbooks'] = Handbook.objects.filter(user=user)
        context['training_materials'] = TrainingMaterial.objects.filter(user=user)
        
        # Add leave types, deductions and allowances for the modal
        context['leave_types'] = LeaveType.objects.filter(is_active=True, user=user)
        context['deductions'] = Deduction.objects.filter(is_active=True, user=user)
        context['allowances'] = Allowance.objects.filter(is_active=True, user=user)
        context['leave_groups'] = LeaveGroup.objects.filter(user=user)
        context['payout_types'] = PayoutType.objects.filter(user=user)
        
        # If token is in kwargs, this is an onboarding form request
        token = kwargs.get('token')
        if token:
            # Try to find invitation by form link
            form_link = f"{self.request.scheme}://{self.request.get_host()}/hr/onboarding/form/{token}/"
            try:
                # Look for invitation with matching token in form_link
                # Filter invitations by the user's companies and the token
                invitations = OnboardingInvitation.objects.filter(form_link__contains=token, company__in=user_companies)
                if invitations.exists():
                    invitation = invitations.first()
                    context['invitation'] = invitation
                    
                    # Check if the form has already been completed/submitted
                    if invitation.is_form_completed:
                        context['is_form_expired'] = True
                        context['completed_at'] = invitation.completed_at
                        # Set template for expired form message
                        self.template_name = 'hr_management/onboarding_form_expired.html'
                    # Check if the invitation has been completed (accepted)
                    elif invitation.status in ['accepted', 'completed']:
                        context['is_onboarding_form'] = True
                        # If using a different template for the form, set it here
                        self.template_name = 'hr_management/onboarding_form.html'
                    else:
                        # Redirect to view offer page if not accepted yet
                        context['redirect_to_offer'] = True
                        context['offer_token'] = token
            except Exception as e:
                logger.error(f"Error retrieving onboarding invitation: {str(e)}", exc_info=True)
        
        # Get all invitations for the "See Invites" tab, filtered by the user's companies
        context['invitations'] = OnboardingInvitation.objects.filter(company__in=user_companies).order_by('-created_at')
        
        return context
        
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        
        # Check if we need to redirect to the offer view page
        if context.get('redirect_to_offer'):
            token = context.get('offer_token')
            return redirect(f"/hr/onboarding/view-offer/{token}/")
            
        return self.render_to_response(context)
    
    def post(self, request, *args, **kwargs):
        """Handle form submission and create employee credentials"""
        try:
            # Get token from URL
            token = kwargs.get('token')
            if not token:
                return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)
                
            # Find invitation by token
            invitation = OnboardingInvitation.objects.get(form_link__contains=token)
            
            # Check if the form has already been completed
            if invitation.is_form_completed:
                return JsonResponse({
                    'success': False, 
                    'message': 'This form has already been submitted and is no longer accessible.'
                }, status=403)
            
            # Log submission attempt
            logger.info(f"Receiving form submission for invitation {invitation.invitation_id} ({invitation.email})")
            
            # Process form data
            form_data = {}
            
            # Process POST data
            for key, value in request.POST.items():
                if key != 'csrfmiddlewaretoken':  # Skip CSRF token
                    form_data[key] = value
            
            # Process FILES data
            files_data = {}
            for key, value in request.FILES.items():
                files_data[key] = value.name
            
            # Add files information to form_data
            form_data['uploaded_files'] = files_data
            
            # Look for JSON structured data
            if 'complete_form_data_json' in form_data:
                try:
                    # Parse the JSON string
                    json_data = json.loads(form_data['complete_form_data_json'])
                    # Include this structured data in our form_data
                    form_data['structured_data'] = json_data
                    logger.info(f"Received structured JSON data for {invitation.email}")
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse complete_form_data_json for {invitation.email}")
            
            # Process JSON serialized personal info if available
            personal_info = {}
            if 'personal_info_json' in form_data:
                try:
                    personal_info = json.loads(form_data['personal_info_json'])
                    form_data['personal_info'] = personal_info
                    logger.info(f"Received personal info JSON for {invitation.email}")
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse personal_info_json for {invitation.email}")
            
            # Process JSON serialized employment details if available
            employment_details = {}
            if 'employment_details_json' in form_data:
                try:
                    employment_details = json.loads(form_data['employment_details_json'])
                    form_data['employment_details'] = employment_details
                    logger.info(f"Received employment details JSON for {invitation.email}")
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse employment_details_json for {invitation.email}")
            
            # Create or update Employee record with form data
            
            
            employee_data = {
                'employee_name': personal_info.get('full_name', invitation.name),
                'employee_email': invitation.email,
                'phone_number': personal_info.get('phone_number'),
                'address': personal_info.get('address'),
                'pan_number': personal_info.get('pan_number'),
                'aadhar_number': personal_info.get('aadhar_number'),
                'dob': personal_info.get('date_of_birth'),
                'qualification': personal_info.get('highest_qualification'),
                'bank_account_holder_name': personal_info.get('bank_account_name'),
                'bank_account_number': personal_info.get('bank_account_number'),
                'bank_name': personal_info.get('bank_name'),
                'branch_name': personal_info.get('branch_name'),
                'bank_ifsc_code': personal_info.get('ifsc_code'),
                'company': invitation.company,
                'is_active': False,
                'is_approved': False
            }
            
            # Find the correct agents.models.Policy instance for employee.department
            # This logic needs to be added once the relationship is clear.
            # For now, setting department to None to avoid errors.
            # Example: if invitation.department and invitation.company:
            # department_policy, _ = Policy.objects.get_or_create(name=invitation.department.name, company=invitation.company) # if Policy has company FK
            # employee_data['department'] = department_policy
            
            employee, created = Employee.objects.update_or_create(
                employee_email=invitation.email,
                defaults=employee_data
            )
            if created:
                logger.info(f"New employee created: {employee}")
            else:
                logger.info(f"Existing employee updated: {employee}")
            # Handle file uploads
            if request.FILES:
                try:
                    # Save profile photo if uploaded
                    if 'profile_photo' in request.FILES:
                        employee.attendance_photo = request.FILES['profile_photo']
                        employee.save()
                    
                    # Handle document uploads
                    from .models import EmployeeDocument
                    
                    for key, file in request.FILES.items():
                        if key != 'profile_photo':  # Already handled above
                            try:
                                # Determine document type from the file key
                                document_type = key.replace('_', ' ').title()
                                    
                                    # Create a document record
                                document = EmployeeDocument(
                                        employee=employee,
                                        document_type=document_type,
                                        document_name=file.name,
                                        file=file,
                                        file_size=str(file.size),
                                    )
                                document.save()
                                logger.info(f"Saved document {document_type} for {invitation.email}")
                            except Exception as file_error:
                                logger.error(f"Error saving document {key}: {str(file_error)}", exc_info=True)
                except Exception as upload_error:
                    logger.error(f"Error handling file uploads: {str(upload_error)}", exc_info=True)
            
            # Store form data in policies for reference
            if invitation.policies is None:
                invitation.policies = {}
            invitation.policies['form_data'] = form_data
            
            # Update invitation status
            invitation.status = 'completed'
            invitation.is_form_completed = True
            invitation.completed_at = timezone.now()
            invitation.save()
            
            # Log successful form submission
            logger.info(f"Form submitted successfully for invitation {invitation.invitation_id} by {invitation.email}")
            
            return JsonResponse({
                'success': True, 
                'message': 'Form submitted successfully. Your application is pending review.'
            })
                
        except OnboardingInvitation.DoesNotExist:
            logger.error(f"Invalid invitation for token: {token}")
            return JsonResponse({'success': False, 'message': 'Invalid invitation'}, status=404)
            
        except Exception as e:
            logger.error(f"Error processing form submission: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    def send_credentials_email(self, invitation, password):
        """Send email with login credentials to the employee"""
        company = invitation.company
        
        # Get site URL from settings or use a default
        site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
        
        # Prepare context data for the email template
        context = {
            'name': invitation.name,
            'email': invitation.email,
            'company_name': company.company_name,
            'department': invitation.department.name if invitation.department else 'Not specified',
            'designation': invitation.designation.name if invitation.designation else 'Not specified',
            'role': invitation.role.name if invitation.role else 'Not specified',
            'login_url': f"{site_url}/hr/employee/login/",
            'username': invitation.email,  # Email is used as username
            'password': password,
            'current_year': timezone.now().year,
        }
        
        # Add company logo if available
        if company.company_logo:
            context['company_logo'] = company.company_logo.url
        
        try:
            # Render email template with context
            html_content = render_to_string('hr_management/email_templates/acceptance_credentials.html', context)
            text_content = strip_tags(html_content)
            
            # Send email
            subject = f"Welcome to {company.company_name} - Your Login Credentials"
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [invitation.email]
            
            send_mail(
                subject=subject,
                message=text_content,
                from_email=from_email,
                recipient_list=recipient_list,
                html_message=html_content,
                fail_silently=False,
            )
            
            logger.info(f"Credentials email sent to {invitation.email}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending credentials email: {str(e)}", exc_info=True)
            return False

def detect_platform(user_agent):
    """Detect platform from user agent string"""
    if not user_agent:
        return "Unknown"
    
    if "Android" in user_agent:
        return "Android"
    elif "iPhone" in user_agent or "iPad" in user_agent or "iPod" in user_agent:
        return "iOS"
    elif "Windows" in user_agent:
        return "Windows"
    elif "Macintosh" in user_agent:
        return "MacOS"
    elif "Linux" in user_agent:
        return "Linux"
    else:
        return "Other"

# Helper function to calculate distance between two geographical coordinates using Haversine formula
def calculate_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [float(lat1), float(lon1), float(lat2), float(lon2)])

    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371000  # Radius of earth in meters
    return c * r

class AttendanceView(TemplateView):
    template_name = 'hr_management/attendance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters from request
        company_filter = self.request.GET.get('company', None)
        search_by = self.request.GET.get('search_by', 'name')
        search_query = self.request.GET.get('search_query', '')
        status_filter = self.request.GET.get('status', None)
        
        # Base queryset for approved employees
        employees = Employee.objects.filter(is_approved=True)
        
        # Apply company filter if specified
        if company_filter:
            employees = employees.filter(company__company_id=company_filter)
        
        # Apply search if provided
        if search_query:
            if search_by == 'name':
                employees = employees.filter(employee_name__icontains=search_query)
            elif search_by == 'mobile':
                employees = employees.filter(phone_number__icontains=search_query)
            elif search_by == 'status':
                if search_query.lower() in ['active', 'inactive']:
                    is_active = search_query.lower() == 'active'
                    employees = employees.filter(is_active=is_active)
        
        # Apply status filter if specified
        if status_filter:
            if status_filter == 'active':
                employees = employees.filter(is_active=True)
            elif status_filter == 'inactive':
                employees = employees.filter(is_active=False)
            elif status_filter == 'on_leave':
                # Get employees currently on leave
                today = timezone.now().date()
                leave_employee_ids = LeaveApplication.objects.filter(
                    status='approved',
                    start_date__lte=today,
                    end_date__gte=today
                ).values_list('employee_id', flat=True)
                employees = employees.filter(employee_id__in=leave_employee_ids)
            elif status_filter == 'unassigned':
                # Unassigned employees - filter based on what makes sense for your application
                # For example, employees without location data
                employees = employees.filter(location__isnull=True)
        
        # Calculate metrics
        today = timezone.now().date()
        
        # 1. Total approved employees
        total_employees = employees.count()
        
        # 2. Active employees (with attendance today)
        active_employee_ids = EmployeeAttendance.objects.filter(
            date=today, 
            status='present'
        ).values_list('employee_id', flat=True)
        active_employees = employees.filter(employee_id__in=active_employee_ids).count()
        
        # 3. Non-attendees (employees who should have attendance but don't)
        non_attendees = total_employees - active_employees
        
        # 4. Employees on leave
        employees_on_leave = LeaveApplication.objects.filter(
            employee_id__in=employees.values_list('employee_id', flat=True),
            status='approved',
            start_date__lte=today,
            end_date__gte=today
        ).count()
        
        # 5. Unassigned employees (without department or designation)
        unassigned_employees = employees.filter(location__isnull=True).count()
        
        # Store metrics in context
        context['metrics'] = {
            'total_employees': total_employees,
            'active_employees': active_employees,
            'non_attendees': non_attendees,
            'employees_on_leave': employees_on_leave,
            'unassigned_employees': unassigned_employees
        }
        
        # Get all employees for display
        context['employees'] = employees
        context['companies'] = Company.objects.all()
        context['qr_codes'] = QRCode.objects.all()
        
        # Pass filters to context for form persistence
        context['filters'] = {
            'company': company_filter,
            'search_by': search_by,
            'search_query': search_query,
            'status': status_filter
        }
        
        return context
        
    def get(self, request, *args, **kwargs):
        # Check if this is an AJAX request for metrics only
        if request.GET.get('fetch_metrics', False) and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            context = self.get_context_data(**kwargs)
            return JsonResponse({'metrics': context['metrics']})
        return super().get(request, *args, **kwargs)


@method_decorator(csrf_exempt, name='dispatch')
class EmployeeStatusToggleAPIView(View):
    def post(self, request, employee_id, *args, **kwargs):
        try:
            # Find the employee - import from correct location
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Get the action (activate or deactivate)
            action = request.POST.get('action', '')
            
            if action == 'activate':
                employee.is_active = True
                employee.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Employee activated successfully',
                    'status': 'active'
                })
            elif action == 'deactivate':
                employee.is_active = False
                employee.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Employee deactivated successfully',
                    'status': 'inactive'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified'
                }, status=400)
                
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Employee not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error toggling employee status: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error processing request: {str(e)}'
            }, status=500)
        

class DeleteCompanyView(View):
    def post(self, request, company_id, *args, **kwargs):
        try:
            company = Company.objects.get(company_id=company_id)
            company.delete()
            return JsonResponse({'success': True, 'message': 'Company deleted successfully'})
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Company not found'}, status=404)
        except Exception as e:
            logger.error(f"Error deleting company: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'Error processing request: {str(e)}'}, status=500)
        
class EditCompanyView(View):
    def post(self, request, company_id, *args, **kwargs):
        try:
            company = Company.objects.get(company_id=company_id)
            company.update(**request.POST)
            return JsonResponse({'success': True, 'message': 'Company updated successfully'})
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Company not found'}, status=404)
        except Exception as e:
            logger.error(f"Error updating company: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'Error processing request: {str(e)}'}, status=500)


class CreateCompanyView(TemplateView):
    template_name = 'hr_management/create_hr_company.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        logger.debug("Rendering create company form")
        return context

    def post(self, request, *args, **kwargs):
        logger.debug("Processing company creation request")
        try:
            # Get user ID from session
            user_session_id = request.session.get('user_session_id')
            if not user_session_id:
                return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)
            
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Invalid session'}, status=401)
                    
            # Get form data
            company_name = request.POST.get('company_name')
            company_logo = request.FILES.get('company_logo')
            company_gst_number = request.POST.get('company_gst_number')
            company_mobile_number = request.POST.get('company_mobile_number')
            company_email = request.POST.get('company_email')
            company_address = request.POST.get('company_address')
            company_identification_number = request.POST.get('company_identification_number')
            company_state = request.POST.get('company_state')
            company_pincode = request.POST.get('company_pincode')

            logger.debug(f"Received company data: name={company_name}, email={company_email}, state={company_state}, pincode={company_pincode}")
            logger.debug(f"Files received: logo={bool(company_logo)}")

            # Create company
            logger.debug("Creating company record")
            company = Company.objects.create(
                company_name=company_name,
                company_logo=company_logo,
                company_gst_number=company_gst_number,
                company_phone=company_mobile_number,
                company_pincode=company_pincode,
                company_email=company_email,
                company_address=company_address,
                company_identification_number=company_identification_number,
                company_state=company_state,
                user=user,
            )

            logger.info(f"Company '{company_name}' created successfully for user {user}")
            return JsonResponse({
                'success': True,
                'redirect_url': '/hr/company/'
            })

        except Exception as e:
            logger.error(f"Error creating company: {str(e)}", exc_info=True)
            return JsonResponse({
                'error': str(e)
            }, status=400)
    
class QRCodeView(View):
    def post(self, request):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            # Determine if we're getting data from JSON or form submission
            if request.content_type and 'application/json' in request.content_type:
                try:
                    data = json.loads(request.body)
                    logger.info(f"Received JSON data: {data}")
                    
                    company_id = data.get('company')
                    locations = data.get('locations', [])
                except json.JSONDecodeError as e:
                    logger.error(f"JSON parsing error: {e}")
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid JSON data'
                    }, status=400)
            else:
                # Handle standard form data
                logger.info(f"Received form data: {request.POST}")
                company_id = request.POST.get('company')
                
                # Get location names and coordinates from form data
                location_names = request.POST.getlist('location_names[]', [])
                coordinates_list = request.POST.getlist('coordinates[]', [])
                
                # Create locations array from form data
                locations = []
                for i in range(len(location_names)):
                    if i < len(coordinates_list):
                        locations.append({
                            'name': location_names[i],
                            'coordinates': coordinates_list[i]
                        })
                
                logger.info(f"Processed form data into locations: {locations}")

            if not company_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Company ID is required'
                }, status=400)
                
            if not locations or len(locations) == 0:
                return JsonResponse({
                    'success': False,
                    'message': 'At least one location with coordinates is required'
                }, status=400)

            # Get company instance
            company = Company.objects.get(company_id=company_id)
            
            created_qr_codes = []

            for location in locations:
                # Extract location name and coordinates
                if isinstance(location, dict):
                    location_name = location.get('name', '').strip()
                    coordinates = location.get('coordinates', '')
                    if isinstance(coordinates, str):
                        coordinates = coordinates.strip()
                else:
                    # Skip if location is not a dict
                    logger.warning(f"Skipping non-dict location: {location}")
                    continue

                if not location_name:
                    logger.warning(f"Skipping location with no name: {location}")
                    continue
                
                # Validate and process coordinates format
                if not coordinates:
                    logger.warning(f"Skipping location with no coordinates: {location}")
                    continue
                
                # Log coordinate format for debugging
                logger.info(f"Processing coordinates format: {type(coordinates)}, value: {coordinates}")
                
                # Handle different coordinate formats
                if isinstance(coordinates, str):
                    # Try to parse coordinates string
                    try:
                        if ',' in coordinates:
                            # Handle "lat,lng" format
                            parts = coordinates.split(',')
                            if len(parts) == 2:
                                try:
                                    lat = float(parts[0].strip())
                                    lng = float(parts[1].strip())
                                    coordinates = {'latitude': lat, 'longitude': lng}
                                    valid_locations_found = True
                                except ValueError:
                                    logger.warning(f"Could not parse coordinates as float: {coordinates}")
                                    continue
                            else:
                                logger.warning(f"Invalid coordinates format (needs exactly one comma): {coordinates}")
                                continue
                        else:
                            try:
                                # Try parsing as JSON string
                                parsed_coords = json.loads(coordinates)
                                if isinstance(parsed_coords, dict):
                                    coordinates = parsed_coords
                                    valid_locations_found = True
                                else:
                                    logger.warning(f"Parsed coordinates not a dict: {parsed_coords}")
                                    continue
                            except json.JSONDecodeError:
                                logger.warning(f"Could not parse coordinates as JSON: {coordinates}")
                                continue
                    except Exception as e:
                        logger.error(f"Error parsing coordinates string: {e}")
                        continue
                elif isinstance(coordinates, dict):
                    # Already in dict format, validate it has required keys
                    if 'latitude' in coordinates and 'longitude' in coordinates:
                        valid_locations_found = True
                    else:
                        logger.warning(f"Coordinates dict missing latitude/longitude: {coordinates}")
                        continue
                else:
                    logger.warning(f"Unsupported coordinates format: {type(coordinates)}")
                    continue

                # Generate QR code and save it
                qr_code_id = uuid.uuid4()
                secret_key = str(uuid.uuid4())

                qr_code = QRCode.objects.create(
                    qr_code_id=qr_code_id,
                    company=company,
                    location_and_coordinates={
                        'location_name': location_name,
                        'coordinates': coordinates
                    },
                    secret_key=secret_key,
                    user=user
                )

                # Create QR code data with attendance URL
                base_url = request.build_absolute_uri('/').rstrip('/')
                attendance_url = f"{base_url}/hr/attend/{qr_code.qr_code_id}/{secret_key}/"
                
                qr_data = {
                    'qr_code_id': str(qr_code.qr_code_id),
                    'company_id': str(company.company_id),
                    'company_name': company.company_name,
                    'location_data': {
                        'name': location_name,
                        'coordinates': coordinates
                    },
                    'secret_key': secret_key,
                    'created_at': qr_code.created_at.isoformat(),
                    'timestamp': timezone.now().isoformat(),
                    'attendance_url': attendance_url
                }

                # Generate and save QR code image
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_H,
                    box_size=10,
                    border=4,
                )
                # Use the URL directly in the QR code instead of the JSON data
                qr.add_data(attendance_url)
                qr.make(fit=True)
                qr_image = qr.make_image(fill_color="black", back_color="white")

                # Save QR code image
                buffer = BytesIO()
                qr_image.save(buffer, format='PNG')
                buffer.seek(0)
                
                filename = f'qr_code_{company.company_name}_{location_name}_{qr_code_id}.png'
                qr_code.qr_code_image.save(filename, ContentFile(buffer.getvalue()), save=True)

                created_qr_codes.append({
                    'id': str(qr_code.qr_code_id),
                    'location': location_name,
                    'image_url': qr_code.qr_code_image.url
                })
            
            if not created_qr_codes or not valid_locations_found:
                return JsonResponse({
                    'success': False,
                    'message': 'No valid QR codes could be generated. Please check your location data.'
                }, status=400)

            return redirect('/hr/company/')

        except Company.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Company not found'
            }, status=404)
            
        except Exception as e:
            logger.error(f"Error generating QR code: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)

class QRCodeDetailsView(View):
    def get(self, request, qr_code_id, *args, **kwargs):
        try:
            # Get the QR code by ID
            qr_code = QRCode.objects.get(qr_code_id=qr_code_id)
            
            # Create response data with QR code details
            response_data = {
                'success': True,
                'qr_code': {
                    'qr_code_id': str(qr_code.qr_code_id),
                    'company_id': str(qr_code.company.company_id),
                    'company_name': qr_code.company.company_name,
                    'location_and_coordinates': qr_code.location_and_coordinates,
                    'secret_key': qr_code.secret_key,
                    'created_at': qr_code.created_at.strftime('%d %b %Y, %I:%M %p'),
                    'qr_code_image': qr_code.qr_code_image.url if qr_code.qr_code_image else None
                }
            }
            
            return JsonResponse(response_data)
            
        except QRCode.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'QR code not found'
            }, status=404)
            
        except Exception as e:
            logger.error(f"Error retrieving QR code details: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error retrieving QR code details: {str(e)}'
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class DeleteQRCodeView(View):
    def post(self, request, qr_code_id, *args, **kwargs):
        try:
            # Get the QR code by ID
            qr_code = QRCode.objects.get(qr_code_id=qr_code_id)
            
            # Store the file path to delete the image after deleting the model
            image_path = qr_code.qr_code_image.path if qr_code.qr_code_image else None
            
            # Delete the QR code
            qr_code.delete()
            
            # Delete the image file from storage if it exists
            if image_path and os.path.isfile(image_path):
                os.remove(image_path)
            
            return JsonResponse({
                'success': True,
                'message': 'QR code deleted successfully'
            })
            
        except QRCode.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'QR code not found'
            }, status=404)
            
        except Exception as e:
            logger.error(f"Error deleting QR code: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error deleting QR code: {str(e)}'
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class QRAttendanceRedirectView(View):
    """View that handles QR code attendance redirects."""
    def get(self, request, qr_code_id, secret_key, *args, **kwargs):
        try:
            # Look up the QR code
            qr_code = get_object_or_404(QRCode, qr_code_id=qr_code_id, secret_key=secret_key)
            
            # Get the location data
            location_data = qr_code.location_and_coordinates or {}
            location_name = location_data.get('location_name', '')
            coordinates = location_data.get('coordinates', {})
            
            # Create URL parameters for the mark attendance page
            params = {
                'qr_code_id': str(qr_code_id),
                'secret_key': secret_key,
                'company_id': str(qr_code.company.company_id),
                'company_name': qr_code.company.company_name,
                'location_name': location_name,
            }
            
            # Add coordinates if available
            if coordinates:
                if isinstance(coordinates, dict):
                    if 'latitude' in coordinates and 'longitude' in coordinates:
                        params['latitude'] = coordinates['latitude']
                        params['longitude'] = coordinates['longitude']
                elif isinstance(coordinates, str):
                    # Handle string format if it's stored that way
                    parts = coordinates.split(',')
                    if len(parts) >= 2:
                        lat_part = parts[0].strip()
                        lng_part = parts[1].strip()
                        if lat_part.startswith('lat:'):
                            params['latitude'] = lat_part[4:]
                        if lng_part.startswith('lng:'):
                            params['longitude'] = lng_part[4:]
            
            # Encode parameters in the URL
            query_params = '&'.join([f"{key}={value}" for key, value in params.items()])
            redirect_url = f"/hr/mark-attendance/?{query_params}&auto_mark=true"
            
            # Redirect to the mark attendance page with parameters
            return redirect(redirect_url)
            
        except Exception as e:
            logger.error(f"Error redirecting to attendance page: {str(e)}", exc_info=True)
            return render(request, 'hr_management/error.html', {
                'error_message': 'Invalid QR code or error processing attendance.'
            })

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeAttendanceView(TemplateView):
    template_name = 'hr_management/mark_attendance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass any QR parameters to the template
        context.update({
            'qr_params': self.request.GET.dict()
        })
        return context

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'send_otp':
                return self.send_otp(request, data)
            elif action == 'verify_otp':
                return self.verify_otp(request, data)
            elif action == 'upload_photo':
                return self.upload_photo(request, data)
            elif action == 'mark_attendance':
                return self.mark_attendance(request, data)
            
            return JsonResponse({'error': 'Invalid action'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
            
    def format_location_data(self, location):
        """Format location data for storage"""
        if not location:
            return "Unknown location"
            
        location_parts = []
        
        # Add location name if available
        if location.get('name'):
            location_parts.append(location.get('name'))
            
        # Add coordinates
        lat = location.get('latitude')
        lng = location.get('longitude')
        if lat is not None and lng is not None:
            location_parts.append(f"({lat}, {lng})")
            
        # Add accuracy if available
        accuracy = location.get('accuracy')
        if accuracy is not None:
            location_parts.append(f"accuracy: {accuracy}m")
            
        # Join all parts
        if location_parts:
            return " - ".join(location_parts)
        else:
            return "Unknown location"

    def send_otp(self, request, data):
        try:
            email = data.get('email')
            
            if not email:
                return JsonResponse({'success': False, 'error': 'Email is required'}, status=400)

            # Add debug logging
            print(f"Attempting to send OTP to email: {email}")

            # Don't modify this line - it's using the hr.models.Employee
            employee = Employee.objects.filter(employee_email=email).first()
            if not employee:
                return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

            # Check if employee is approved
            if not employee.is_approved:
                return JsonResponse({'success': False, 'error': 'Your account is pending approval'}, status=403)

            # Generate a 6-digit OTP
            otp = ''.join(random.choices('0123456789', k=6))
            
            # Store OTP in session
            request.session['attendance_otp'] = {
                'code': otp,
                'email': email,
                'timestamp': timezone.now().isoformat()
            }

            try:
                # Add more descriptive email subject and body
                subject = f'{employee.company.company_name} - Attendance OTP'
                message = (
                    f'Hello {employee.employee_name},\n\n'
                    f'Your OTP for attendance verification is: {otp}\n'
                    f'This OTP will expire in 5 minutes.\n\n'
                    f'If you did not request this OTP, please ignore this email.'
                )
                
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                print(f"OTP sent successfully to {email}")
                return JsonResponse({'success': True, 'message': 'OTP sent successfully'})
            except Exception as e:
                print(f"Failed to send OTP: {str(e)}")
                return JsonResponse({
                    'success': False, 
                    'error': f'Failed to send OTP email: {str(e)}'
                }, status=500)

        except Exception as e:
            print(f"Error in send_otp: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def verify_otp(self, request, data):
        try:
            entered_otp = data.get('otp')
            email = data.get('email')

            if not entered_otp or not email:
                return JsonResponse({'success': False, 'error': 'OTP and email are required'}, status=400)

            stored_otp = request.session.get('attendance_otp', {})
            
            if not stored_otp or stored_otp.get('email') != email:
                return JsonResponse({'success': False, 'error': 'Invalid session'}, status=400)

            otp_timestamp = datetime.fromisoformat(stored_otp['timestamp'])
            if timezone.now() - otp_timestamp > timedelta(minutes=5):
                return JsonResponse({'success': False, 'error': 'OTP expired'}, status=400)

            if entered_otp != stored_otp['code']:
                return JsonResponse({'success': False, 'error': 'Invalid OTP'}, status=400)

            # Generate a unique device ID for trusted device
            device_id = str(uuid.uuid4())
            
            # Get employee for future use
            employee = Employee.objects.filter(employee_email=email).first()
            if not employee:
                return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
                
            # Get device info from user agent
            user_agent = request.META.get('HTTP_USER_AGENT', '')
            
            # Store the trusted device in the database
            try:
                from .models import TrustedDevice
                
                # Create a new trusted device record
                trusted_device = TrustedDevice.objects.create(
                    device_id=device_id,
                    employee=employee,
                    browser_info=user_agent[:255] if user_agent else None,
                    platform=detect_platform(user_agent),
                )
                
                # Log the creation
                print(f"Created trusted device: {trusted_device}")
                
            except Exception as e:
                # Log the error but continue - we'll still use client-side storage
                print(f"Error creating trusted device record: {str(e)}")

            # Create a device trust token
            return JsonResponse({
                'success': True,
                'message': 'OTP verified successfully',
                'device_id': device_id,
                'employee_name': employee.employee_name
            })

        except Exception as e:
            print(f"Error in verify_otp: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def upload_photo(self, request, data):
        try:
            print(f"Starting photo upload for request: {request}")
            photo_data = data.get('attendance_image')  # From frontend
            device_id = data.get('device_id')
            email = data.get('email') or request.session.get('attendance_otp', {}).get('email')  # Get email from request or session

            if not photo_data or not device_id or not email:
                print(f"Missing data - photo_data: {bool(photo_data)}, device_id: {bool(device_id)}, email: {bool(email)}")
                return JsonResponse({'success': False, 'error': 'Missing required data'}, status=400)

            print(f"Processing photo for device ID: {device_id} and email: {email}")
            
            # Handle base64 data with or without prefix
            if 'base64,' in photo_data:
                format, imgstr = photo_data.split('base64,')
                ext = format.split('/')[-1].split(';')[0]
            else:
                imgstr = photo_data
                ext = 'jpg'

            # Clean base64 string
            imgstr = imgstr.strip()
            
            try:
                # Convert base64 to file
                photo = ContentFile(base64.b64decode(imgstr), name=f'attendance_{device_id}.{ext}')
                
                # Save to Employee model using the correct field name
                # Don't modify this line - it's using the hr.models.Employee
                employee = Employee.objects.get(employee_email=email)
                employee.attendance_photo = photo  # Using attendance_photo to match model field
                employee.save()
                
                print(f"Successfully saved attendance photo for employee: {employee.employee_email}")
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Photo uploaded successfully',
                    'photo_url': employee.attendance_photo.url if employee.attendance_photo else None
                })
                
            except Employee.DoesNotExist:
                print(f"Employee not found with email: {email}")
                return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
            except Exception as e:
                print(f"Error processing photo data: {str(e)}")
                return JsonResponse({'success': False, 'error': 'Invalid photo data format'}, status=400)

        except Exception as e:
            print(f"Error uploading photo: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def mark_attendance(self, request, data):
        try:
            # Get trusted device info
            device_id = data.get('device_id')
            email = data.get('email')
            qr_data = data.get('qr_data')
            current_location = data.get('location')

            if not all([qr_data, current_location, device_id, email]):
                return JsonResponse({'success': False, 'error': 'Missing required data'}, status=400)

            # Log the received data for debugging
            print(f"QR data received: {qr_data}")
            print(f"Location data received: {current_location}")

            # Get employee
            try:
                # Don't modify this line - it's using the hr.models.Employee
                employee = Employee.objects.get(employee_email=email)
                
                # Check if employee is approved
                if not employee.is_approved:
                    return JsonResponse({'success': False, 'error': 'Your account is pending approval'}, status=403)
                
                # Verify the trusted device
                from .models import TrustedDevice
                try:
                    # Look up the trusted device in our database
                    trusted_device = TrustedDevice.objects.get(
                        device_id=device_id,
                        employee=employee,
                        is_active=True
                    )
                    
                    # Update last used timestamp
                    trusted_device.save()  # This will update the auto_now last_used field
                    
                except TrustedDevice.DoesNotExist:
                    # Device not found in our database, but we'll still allow marking attendance
                    # since we have client-side verification as a backup
                    print(f"Warning: Trusted device {device_id} not found in database for employee {email}")
                
                # Check if attendance was marked today
                now = timezone.now()
                today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                
                if employee.attendance_status == 'completed':
                    return JsonResponse({
                        'success': False,
                        'error': 'You have already completed attendance for today'
                    }, status=400)

                # If last attendance was marked within 15 minutes
                if (employee.last_attendance_time and 
                    employee.attendance_status == 'marked' and 
                    (now - employee.last_attendance_time).total_seconds() < 900):  # 15 minutes = 900 seconds
                    return JsonResponse({
                        'success': False,
                        'error': 'You have already marked attendance. Please wait 15 minutes before unmarking.'
                    }, status=400)

                # More flexible QR code data handling
                qr_coordinates = None
                # Try different possible structures for QR code data
                if 'location_data' in qr_data and 'coordinates' in qr_data.get('location_data', {}):
                    qr_coordinates = qr_data['location_data']['coordinates']
                elif 'coordinates' in qr_data:
                    qr_coordinates = qr_data['coordinates']
                elif 'latitude' in qr_data and 'longitude' in qr_data:
                    qr_coordinates = {'latitude': qr_data['latitude'], 'longitude': qr_data['longitude']}
                
                # If we have QR coordinates, proceed with distance verification
                if qr_coordinates and 'latitude' in qr_coordinates and 'longitude' in qr_coordinates:
                    qr_lat = qr_coordinates['latitude']
                    qr_lon = qr_coordinates['longitude']
                    
                    device_lat = current_location.get('latitude')
                    device_lon = current_location.get('longitude')
                    
                    if device_lat is not None and device_lon is not None:
                        # Calculate distance between QR code location and device location
                        try:
                            distance = calculate_distance(qr_lat, qr_lon, device_lat, device_lon)
                            
                            # Check if distance is greater than 20 meters
                            if distance > 20:
                                return JsonResponse({
                                    'success': False,
                                    'error': f'You are too far from the attendance location. Maximum allowed distance is 20 meters, you are {int(distance)} meters away.'
                                }, status=400)
                                
                            print(f"Device is {distance:.2f} meters from QR code location")
                            
                            # Get location accuracy for logging
                            accuracy = current_location.get('accuracy', 'unknown')
                            print(f"Location accuracy: {accuracy} meters")
                        except Exception as e:
                            print(f"Error calculating distance: {e}")
                            # Continue without strict distance check if there's an error
                    else:
                        print("Device location coordinates missing")
                        # For demo purposes, we'll continue without strict location check
                        # In production, you might want to enforce this
                        # return JsonResponse({
                        #     'success': False,
                        #     'error': 'Unable to determine your current location. Please enable location services.'
                        # }, status=400)
                else:
                    print("QR coordinates missing or in unexpected format")
                    # For demo purposes, we'll continue without strict location check
                
                # Update attendance based on current status
                if employee.attendance_status == 'not_marked':
                    employee.number_of_days_attended += 1
                    employee.attendance_status = 'marked'
                    message = 'Attendance marked successfully!'
                    
                    # Format location data for storage
                    location_string = self.format_location_data(current_location)
                    
                    # Create an attendance record
                    today = timezone.now().date()
                    attendance, created = EmployeeAttendance.objects.get_or_create(
                        employee=employee,
                        date=today,
                        defaults={
                            'status': 'present',
                            'check_in_time': now,
                            'location': location_string
                        }
                    )
                    
                    if not created:
                        attendance.check_in_time = now
                        attendance.save()
                        
                elif employee.attendance_status == 'marked':
                    employee.attendance_status = 'completed'
                    message = 'Check-out recorded successfully. You cannot mark attendance again today.'
                    
                    # Format location data for storage
                    location_string = self.format_location_data(current_location)
                    
                    # Update attendance record with check-out time
                    today = timezone.now().date()
                    try:
                        attendance = EmployeeAttendance.objects.get(
                            employee=employee,
                            date=today
                        )
                        attendance.check_out_time = now
                        attendance.check_out_location = location_string
                        attendance.save()
                    except EmployeeAttendance.DoesNotExist:
                        # Create a new record if one doesn't exist
                        EmployeeAttendance.objects.create(
                            employee=employee,
                            date=today,
                            status='present',
                            check_in_time=now,
                            check_out_time=now,
                            location=location_string
                        )

                employee.last_attendance_time = now
                employee.save()

                return JsonResponse({
                    'success': True,
                    'message': message,
                    'status': employee.attendance_status
                })
                
            except Employee.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

        except Exception as e:
            print(f"Error marking attendance: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

class CreateFolderView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Get form data
            name = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            
            # Debug request
            print(f"Received folder creation request - Title: {name}, Description: {description}")
            print(f"POST data: {request.POST}")
            print(f"FILES data: {request.FILES}")
            
            # Validate input
            if not name:
                return JsonResponse({'success': False, 'error': 'Folder name is required'})
            
            # Generate unique folder ID
            folder_id = str(uuid.uuid4())
            
            # Handle file upload
            logo = request.FILES.get('logo')
            logo_path = None
            
            if logo:
                try:
                    # Validate file type
                    if not logo.content_type.startswith('image/'):
                        return JsonResponse({
                            'success': False, 
                            'error': 'Invalid file type. Please upload an image.'
                        })

                    # Generate unique filename
                    file_ext = os.path.splitext(logo.name)[1]
                    unique_filename = f"folder_logos/{folder_id}{file_ext}"
                    
                    # Save file using default storage
                    logo_path = default_storage.save(
                        unique_filename,
                        ContentFile(logo.read())
                    )
                    print(f"Saved logo to {logo_path}")
                except Exception as e:
                    logger.error(f"Error processing logo: {str(e)}")
                    return JsonResponse({
                        'success': False,
                        'error': f'Error processing logo: {str(e)}'
                    })

            try:
                # Get user from session
                user = None
                user_session_id = request.session.get('user_session_id')
                if user_session_id:
                    try:
                        user_session = UserSession.objects.get(id=user_session_id)
                        user = user_session.user
                        print(f"Found user: {user.user_id}")
                    except UserSession.DoesNotExist:
                        print("User not found")
                        pass
                else:
                    print("No user ID in session")
                
                # Create folder object
                folder = Folder.objects.create(
                    folder_id=folder_id,
                    name=name,
                    description=description,
                    logo=logo_path,
                    created_at=timezone.now(),
                    user=get_user_from_session(self.request)
                )
                
                # Associate with user if available
                if user:
                    folder.user = user
                    folder.save()
                    print(f"Associated folder with user: {user.user_id}")
                
                print(f"Created folder: {folder.folder_id}, {folder.name}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Folder created successfully',
                    'folder_id': str(folder.folder_id),
                    'name': folder.name,
                    'description': folder.description,
                    'logo_url': folder.logo.url if folder.logo else None
                })
                
            except Exception as e:
                # Clean up uploaded file if folder creation fails
                if logo_path:
                    default_storage.delete(logo_path)
                logger.error(f"Error creating folder: {str(e)}")
                raise e

        except Exception as e:
            logger.error(f"Error in CreateFolderView: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error creating folder: {str(e)}'
            })

    def get(self, request, *args, **kwargs):
        try:
            folders = Folder.objects.all().order_by('-created_at')
            context = {
                'folders': folders,
                'app_name': 'Folders'
            }
            return render(request, 'hr_management/creation.html', context)
        except Exception as e:
            logger.error(f"Error in get folders: {str(e)}")
            return render(request, 'hr_management/creation.html', {'error': str(e)})
        
        
class FolderView(TemplateView):
    template_name = 'hr_management/folder.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        folder_id = kwargs.get('folder_id')
        folder = Folder.objects.get(folder_id=folder_id)
        context['app_name'] = folder.name
        context['folder'] = folder
        return context

    def post(self, request, *args, **kwargs):
        try:
            folder_id = kwargs.get('folder_id')
            folder = Folder.objects.get(folder_id=folder_id)

            data = json.loads(request.body)
            title = data.get('title')
            content = data.get('content')

            if not title or not content:
                return JsonResponse({
                    'success': False,
                    'error': 'Title and content are required'
                }, status=400)

            # Initialize json_data if None
            if folder.json_data is None:
                folder.json_data = {}

            # Add new key-value pair
            folder.json_data[title] = content
            folder.save()

            return JsonResponse({
                'success': True,
                'message': 'Data added successfully',
                'title': title,
                'content': content
            })

        except Folder.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Folder not found'
            }, status=404)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error adding folder data: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error adding data: {str(e)}'
            }, status=500)

    def delete(self, request, *args, **kwargs):
        try:
            folder_id = kwargs.get('folder_id')
            folder = Folder.objects.get(folder_id=folder_id)

            data = json.loads(request.body)
            title = data.get('title')

            if not title:
                return JsonResponse({
                    'success': False,
                    'error': 'Title is required for deletion'
                }, status=400)

            if folder.json_data and title in folder.json_data:
                del folder.json_data[title]
                folder.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Data deleted successfully'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Title not found in folder data'
                }, status=404)

        except Folder.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Folder not found'
            }, status=404)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error deleting folder data: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error deleting data: {str(e)}'
            }, status=500)

    def put(self, request, *args, **kwargs):
        try:
            folder_id = kwargs.get('folder_id')
            folder = Folder.objects.get(folder_id=folder_id)

            data = json.loads(request.body)
            original_title = data.get('originalTitle')
            new_title = data.get('newTitle')
            new_content = data.get('newContent')

            if not original_title or not new_title or not new_content:
                return JsonResponse({
                    'success': False,
                    'error': 'Original title, new title and new content are required'
                }, status=400)

            if folder.json_data and original_title in folder.json_data:
                # Delete old key and add new key-value pair
                del folder.json_data[original_title]
                folder.json_data[new_title] = new_content
                folder.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Data updated successfully',
                    'title': new_title,
                    'content': new_content
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Title not found in folder data'
                }, status=404)

        except Folder.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Folder not found'
            }, status=404)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error updating folder data: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error updating data: {str(e)}'
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Handle form data with potential file uploads
            if 'multipart/form-data' in request.content_type:
                # Get form data
                company_id = request.POST.get('company')
                name = request.POST.get('name')
                email = request.POST.get('email')
                department_id = request.POST.get('department')
                designation_id = request.POST.get('designation')
                role_id = request.POST.get('role')
                offer_letter_id = request.POST.get('offer_letter')
                hiring_agreement_id = request.POST.get('hiring_agreement')
                handbook_id = request.POST.get('handbook')
                hr_policies_id = request.POST.get('hr_policies')
                training_material_id = request.POST.get('training_material')
                policies_json = request.POST.get('policies')
                
                # Parse policies JSON if present
                policies = json.loads(policies_json) if policies_json else []
                
                # Log information about the document selections for debugging
                logger.info(f"Processing onboarding invitation for {name} ({email})")
                logger.info(f"Selected documents: offer_letter={offer_letter_id}, hiring_agreement={hiring_agreement_id}, handbook={handbook_id}, hr_policies={hr_policies_id}, training_material={training_material_id}")
                
                # Get photo if uploaded
                photo = request.FILES.get('photo')
            # Parse the JSON data if content type is application/json
            elif request.content_type == 'application/json':
                data = json.loads(request.body)
                
                # Get form data
                company_id = data.get('company')
                name = data.get('name')
                email = data.get('email')
                department_id = data.get('department')
                designation_id = data.get('designation')
                role_id = data.get('role')
                offer_letter_id = data.get('offer_letter')
                hiring_agreement_id = data.get('hiring_agreement')
                handbook_id = data.get('handbook')
                hr_policies_id = data.get('hr_policies')
                training_material_id = data.get('training_material')
                policies = data.get('policies', [])
                
                # No photo in JSON request
                photo = None
            else:
                # Regular form submission
                data = request.POST
                
                # Get form data
                company_id = data.get('company')
                name = data.get('name')
                email = data.get('email')
                department_id = data.get('department')
                designation_id = data.get('designation')
                role_id = data.get('role')
                offer_letter_id = data.get('offer_letter')
                hiring_agreement_id = data.get('hiring_agreement')
                handbook_id = data.get('handbook')
                hr_policies_id = data.get('hr_policies')
                training_material_id = data.get('training_material')
                policies = data.get('policies', [])
                
                # No photo in regular form post
                photo = None
            
            # Validate required fields
            if not company_id or not name or not email:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields: company, name, email'
                }, status=400)
            
            # Get company instance
            try:
                company = Company.objects.get(company_id=company_id)
            except Company.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Company not found'
                }, status=404)
            
            # Get department, designation, role instances if provided
            department = None
            designation = None
            role = None
            
            if department_id:
                try:
                    department = Department.objects.get(department_id=department_id)
                except Department.DoesNotExist:
                    logger.warning(f"Department with ID {department_id} not found")
            
            if designation_id:
                try:
                    designation = Designation.objects.get(designation_id=designation_id)
                except Designation.DoesNotExist:
                    logger.warning(f"Designation with ID {designation_id} not found")
            
            if role_id:
                try:
                    role = Role.objects.get(role_id=role_id)
                except Role.DoesNotExist:
                    logger.warning(f"Role with ID {role_id} not found")
            
            # Get offer letter instance if provided
            offer_letter = None
            if offer_letter_id:
                try:
                    offer_letter = OfferLetter.objects.get(offer_letter_id=offer_letter_id)
                except OfferLetter.DoesNotExist:
                    logger.warning(f"Offer letter with ID {offer_letter_id} not found")
            
            # Get hiring agreement instance if provided
            hiring_agreement = None
            if hiring_agreement_id:
                try:
                    hiring_agreement = HiringAgreement.objects.get(hiring_agreement_id=hiring_agreement_id)
                except HiringAgreement.DoesNotExist:
                    logger.warning(f"Hiring agreement with ID {hiring_agreement_id} not found")
            
            # Get handbook instance if provided
            handbook = None
            if handbook_id:
                try:
                    handbook = Handbook.objects.get(handbook_id=handbook_id)
                except Handbook.DoesNotExist:
                    logger.warning(f"Handbook with ID {handbook_id} not found")
            
            # Get HR policies instance if provided
            hr_policies = None
            if hr_policies_id:
                try:
                    hr_policies = TandC.objects.get(tandc_id=hr_policies_id)
                except TandC.DoesNotExist:
                    logger.warning(f"HR policies with ID {hr_policies_id} not found")
            
            # Get training material instance if provided
            training_material = None
            if training_material_id:
                try:
                    training_material = TrainingMaterial.objects.get(training_material_id=training_material_id)
                except TrainingMaterial.DoesNotExist:
                    logger.warning(f"Training material with ID {training_material_id} not found")
            
            # Generate a unique form link with token
            token = str(uuid.uuid4())
            form_link = f"{request.scheme}://{request.get_host()}/hr/onboarding/form/{token}/"
            
            # Get additional documents if provided
            additional_documents_json = request.POST.get('additional_documents')
            additional_documents = None
            if additional_documents_json:
                try:
                    additional_documents = json.loads(additional_documents_json)
                except json.JSONDecodeError:
                    logger.error(f"Invalid JSON in additional_documents: {additional_documents_json}")
                    
            # Create invitation object
            invitation = OnboardingInvitation.objects.create(
                company=company,
                name=name,
                email=email,
                department=department,
                designation=designation,
                role=role,
                offer_letter_template=offer_letter,
                hiring_agreement_template=hiring_agreement,
                handbook_template=handbook,
                hr_policies_template=hr_policies,
                training_material_template=training_material,
                form_link=form_link,
                policies=policies,
                additional_documents=additional_documents,
                photo=photo,
                user=get_user_from_session(self.request)
            )
            
            # Log successful creation with document details
            logger.info(f"Created invitation {invitation.invitation_id} with documents:")
            logger.info(f"  - Offer Letter: {offer_letter.name if offer_letter else 'None'}")
            logger.info(f"  - Hiring Agreement: {hiring_agreement.name if hiring_agreement else 'None'}")
            logger.info(f"  - Handbook: {handbook.name if handbook else 'None'}")
            logger.info(f"  - HR Policies: {hr_policies.name if hr_policies else 'None'}")
            logger.info(f"  - Training Material: {training_material.name if training_material else 'None'}")
            
            # Send email
            logger.info(f"Attempting to send invitation email to {email}")
            email_sent = self.send_invitation_email(invitation)
            
            # Update invitation status
            if email_sent:
                invitation.status = 'sent'
                invitation.sent_at = timezone.now()
                invitation.save()
                logger.info(f"Invitation status updated to 'sent' for {email}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Invitation sent successfully',
                    'invitation_id': str(invitation.invitation_id)
                })
            else:
                # Email failed, but invitation was created
                logger.warning(f"Invitation created but email failed for {email}")
                return JsonResponse({
                    'success': True,
                    'warning': 'Invitation created but email could not be sent',
                    'invitation_id': str(invitation.invitation_id)
                })
            
        except Exception as e:
            logger.error(f"Error sending onboarding invitation: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)
    
    def send_invitation_email(self, invitation):
        """Send onboarding invitation email"""
        try:
            company = invitation.company
            
            # Get site URL from settings or use a default
            site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
            
            # Generate a random password for the user
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
            
            # Extract token from form_link
            token = invitation.form_link.split('/')[-2]
            
            # Create view offer link
            view_offer_link = f"{site_url}/hr/onboarding/view-offer/{token}/"
            
            # Prepare context data for the email template
            context = {
                'name': invitation.name,
                'email': invitation.email,
                'company_name': company.company_name,
                'department': invitation.department.name if invitation.department else 'Not specified',
                'designation': invitation.designation.name if invitation.designation else 'Not specified',
                'role': invitation.role.name if invitation.role else 'Not specified',
                'login_url': f"{site_url}/hr/employee/login/",
                'form_link': invitation.form_link,
                'view_offer_link': view_offer_link,
                'username': invitation.email,  # Email is used as username
                'password': password,
                'current_year': timezone.now().year,
            }
            
            # Add company logo if available
            if company.company_logo:
                context['company_logo'] = company.company_logo.url
            
            # Log that we're about to render the template
            logger.info(f"Preparing to send invitation email to {invitation.email}")
            
            # Check if template exists
            template_path = 'hr_management/email_templates/onboarding_invitation.html'
            
            # Render email template with context
            try:
                html_content = render_to_string(template_path, context)
                text_content = strip_tags(html_content)
                
                # Log email content for debugging
                logger.debug(f"Email HTML content: {html_content[:100]}...")
                
                # Send email
                subject = f"Welcome to {company.company_name} - Onboarding Invitation"
                from_email = settings.DEFAULT_FROM_EMAIL
                
                if not from_email:
                    logger.error("DEFAULT_FROM_EMAIL is not set in settings")
                    from_email = "noreply@1matrix.io"
                
                recipient_list = [invitation.email]
                
                logger.info(f"Sending email to {invitation.email} with subject '{subject}' from {from_email}")
                
                email = EmailMessage(
                    subject,
                    html_content,
                    from_email,
                    recipient_list,
                )
                email.content_subtype = "html"  # Main content is HTML
                
                # Attach offer letter
                if invitation.offer_letter_template and invitation.offer_letter_template.content:
                    pdf_content = generate_pdf_from_html(invitation.offer_letter_template.content)
                    if pdf_content:
                        email.attach(f"{invitation.offer_letter_template.name}.pdf", pdf_content, 'application/pdf')

                # Attach hiring agreement
                if invitation.hiring_agreement_template and invitation.hiring_agreement_template.content:
                    pdf_content = generate_pdf_from_html(invitation.hiring_agreement_template.content)
                    if pdf_content:
                        email.attach(f"{invitation.hiring_agreement_template.name}.pdf", pdf_content, 'application/pdf')

                # Attach handbook
                if invitation.handbook_template and invitation.handbook_template.content:
                    pdf_content = generate_pdf_from_html(invitation.handbook_template.content)
                    if pdf_content:
                        email.attach(f"{invitation.handbook_template.name}.pdf", pdf_content, 'application/pdf')

                # Attach HR policies
                if invitation.hr_policies_template and invitation.hr_policies_template.description:
                    pdf_content = generate_pdf_from_html(invitation.hr_policies_template.description)
                    if pdf_content:
                        email.attach(f"{invitation.hr_policies_template.name}.pdf", pdf_content, 'application/pdf')

                # Attach training material
                if invitation.training_material_template and invitation.training_material_template.content:
                    pdf_content = generate_pdf_from_html(invitation.training_material_template.content)
                    if pdf_content:
                        email.attach(f"{invitation.training_material_template.name}.pdf", pdf_content, 'application/pdf')

                # Attach additional documents
                if invitation.additional_documents:
                    for doc_info in invitation.additional_documents:
                        try:
                            doc_id = doc_info.get('id')
                            if not doc_id:
                                continue
                            
                            doc = EmployeeDocument.objects.get(document_id=doc_id)
                            
                            if doc.file:
                                with doc.file.open('rb') as f:
                                    file_content = f.read()
                                
                                import mimetypes
                                import os

                                file_name = os.path.basename(doc.file.name)
                                content_type, _ = mimetypes.guess_type(file_name)
                                if content_type is None:
                                    content_type = 'application/octet-stream'
                                
                                email.attach(file_name, file_content, content_type)

                        except EmployeeDocument.DoesNotExist:
                            logger.warning(f"Additional document with ID {doc_id} not found")
                        except Exception as e:
                            logger.error(f"Error attaching document with ID {doc_id}: {e}", exc_info=True)

                # Send the email
                email.send(fail_silently=False)
                logger.info(f"Successfully sent invitation email to {invitation.email}")
                return True
                
            except Exception as template_error:
                logger.error(f"Error rendering or sending email template: {str(template_error)}", exc_info=True)
                return False
                
        except Exception as e:
            logger.error(f"Error in send_invitation_email: {str(e)}", exc_info=True)
            return False

def generate_pdf_from_html(html_content):
    """
    Generates a PDF from HTML content.
    """
    try:
        result = BytesIO()
        # The HTML content might be a template path or raw HTML
        template = get_template("hr_management/pdf_template.html")
        html = template.render({'content': html_content})
        
        # Create PDF
        pisa_status = pisa.CreatePDF(
            html,                # the HTML to convert
            dest=result           # file-like object to receive PDF
        )
        
        if pisa_status.err:
            logger.error(f"PDF generation error: {pisa_status.err}")
            return None
        
        result.seek(0)
        return result.getvalue()
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
        return None

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationStatusView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        """Update invitation status"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Parse request data
            try:
                data = json.loads(request.body)
                status = data.get('status')
                completion_data = data.get('completion_data', {})
            except (ValueError, json.JSONDecodeError):
                # Try to get from POST data if not JSON
                status = request.POST.get('status')
                completion_data = {}
            
            # Check status
            if status not in ['completed', 'rejected', 'accepted']:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid status provided'
                }, status=400)
            
            # Update invitation
            invitation.status = status
            
            # Set appropriate timestamp based on status
            if status == 'completed':
                invitation.completed_at = timezone.now()
                invitation.is_form_completed = True
            elif status == 'rejected':
                invitation.rejected_at = timezone.now()
            elif status == 'accepted':
                invitation.accepted_at = timezone.now()
            
            # Handle additional data if provided
            if status == 'rejected' and completion_data.get('rejection_reason'):
                invitation.rejection_reason = completion_data.get('rejection_reason')
            
            # Save the invitation
            invitation.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Invitation status updated to {status}'
            })
            
        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error updating invitation status: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)

class DepartmentTemplateView(TemplateView):
    template_name = 'hr_management/templates/department.html'
    
    def get_context_data(self, **kwargs):
        user_session_id = self.request.session.get('user_session_id')
        if not user_session_id:
            return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)
            
        try:
            user_session = UserSession.objects.get(id=user_session_id)
            user = user_session.user
        except UserSession.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid session'}, status=401)
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.filter(user=user)
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            department_name = request.POST.get('department_name')
            if department_name:
                Department.objects.create(name=department_name, user=get_user_from_session(self.request))
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Department name is required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class DesignationTemplateView(TemplateView):
    template_name = 'hr_management/templates/designation.html'
    
    def get_context_data(self, **kwargs):
        user_session_id = self.request.session.get('user_session_id')
        if not user_session_id:
            return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)
            
        try:
            user_session = UserSession.objects.get(id=user_session_id)
            user = user_session.user
        except UserSession.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid session'}, status=401)
        context = super().get_context_data(**kwargs)
        context['designations'] = Designation.objects.filter(user=user)
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            designation_name = request.POST.get('designation_name')
            if designation_name:
                Designation.objects.create(name=designation_name, user=get_user_from_session(self.request))
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Designation name is required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class RoleTemplateView(TemplateView):
    template_name = 'hr_management/templates/role.html'
    
    def get_context_data(self, **kwargs):
        user_session_id = self.request.session.get('user_session_id')
        if not user_session_id:
            return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)
        try:
            user_session = UserSession.objects.get(id=user_session_id)
            user = user_session.user
        except UserSession.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid session'}, status=401)
        context = super().get_context_data(**kwargs)
        context['roles'] = Role.objects.filter(user=user)
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            role_name = request.POST.get('role_name')
            if role_name:
                Role.objects.create(name=role_name, user=get_user_from_session(self.request))
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Role name is required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class OfferLetterTemplateView(TemplateView):
    template_name = 'hr_management/templates/offerletter.html'
    
    def get_context_data(self, **kwargs):
        user_session_id = self.request.session.get('user_session_id')
        if not user_session_id:
            return JsonResponse({'success': False, 'message': 'User not authenticated'}, status=401)
        try:
            user_session = UserSession.objects.get(id=user_session_id)
            user = user_session.user
        except UserSession.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid session'}, status=401)
        context = super().get_context_data(**kwargs)
        context['offerletters'] = OfferLetter.objects.filter(user=get_user_from_session(self.request))
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            name = request.POST.get('offer_letter_name')
            content = request.POST.get('offer_letter_content')
            if name and content:
                OfferLetter.objects.create(name=name, content=content, user=get_user_from_session(self.request))
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class PolicyTemplateView(TemplateView):
    template_name = 'hr_management/templates/policy.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['policies'] = TandC.objects.filter(user=get_user_from_session(self.request))
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            name = request.POST.get('policy_name')
            description = request.POST.get('policy_description')
            if name and description:
                TandC.objects.create(name=name, description=description, user=get_user_from_session(self.request))
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Name and description are required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class HiringAgreementTemplateView(TemplateView):
    template_name = 'hr_management/templates/hiring_agreement.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = None
        user_session_id = self.request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None
        
        if user:
            context['agreements'] = HiringAgreement.objects.filter(user=user)
        else:
            context['agreements'] = HiringAgreement.objects.all()
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            name = request.POST.get('agreement_name')
            content = request.POST.get('agreement_content')
            if name and content:
                HiringAgreement.objects.create(name=name, content=content, user=user)
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class HandbookTemplateView(TemplateView):
    template_name = 'hr_management/templates/handbook.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = None
        user_session_id = self.request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None
        
        if user:
            context['handbooks'] = Handbook.objects.filter(user=user)
        else:
            context['handbooks'] = Handbook.objects.all()
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            name = request.POST.get('handbook_name')
            content = request.POST.get('handbook_content')
            if name and content:
                Handbook.objects.create(name=name, content=content, user=user)
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class TrainingMaterialTemplateView(TemplateView):
    template_name = 'hr_management/templates/training_material.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = None
        user_session_id = self.request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None
        
        if user:
            context['materials'] = TrainingMaterial.objects.filter(user=user)
        else:
            context['materials'] = TrainingMaterial.objects.all()
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            name = request.POST.get('material_name')
            content = request.POST.get('material_content')
            if name and content:
                TrainingMaterial.objects.create(name=name, content=content, user=user)
                return JsonResponse({'success': True})
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class OfferLetterPreviewView(View):
    def get(self, request, template_id):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            if user:
                template = OfferLetter.objects.get(offer_letter_id=template_id, user=user)
            else:
                template = OfferLetter.objects.get(offer_letter_id=template_id)
            return JsonResponse({
                'success': True,
                'content': template.content
            })
        except OfferLetter.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

class HiringAgreementPreviewView(View):
    def get(self, request, template_id):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            if user:
                template = HiringAgreement.objects.get(hiring_agreement_id=template_id, user=user)
            else:
                template = HiringAgreement.objects.get(hiring_agreement_id=template_id)
            return JsonResponse({
                'success': True,
                'content': template.content
            })
        except HiringAgreement.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

class HandbookPreviewView(View):
    def get(self, request, template_id):
        try:
            user = None
            user_session_id = request.session.get('user_session_id')
            if user_session_id:
                try:
                    user_session = UserSession.objects.get(id=user_session_id)
                    user = user_session.user
                except UserSession.DoesNotExist:
                    pass
            
            if user:
                template = Handbook.objects.get(handbook_id=template_id, user=user)
            else:
                template = Handbook.objects.get(handbook_id=template_id)
            return JsonResponse({
                'success': True,
                'content': template.content
            })
        except Handbook.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationDetailView(View):
    def get(self, request, invitation_id, *args, **kwargs):
        """View an invitation and provide accept/reject options"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Check if JSON response is requested - check URL parameters first, then headers
            wants_json = (
                request.GET.get('format') == 'json' or 
                request.GET.get('v') == '2' or 
                'application/json' in request.META.get('HTTP_ACCEPT', '') or
                request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            )
            
            # For HTML template rendering
            if not wants_json:
                context = {
                    'invitation': invitation,
                    'current_year': timezone.now().year,
                }
                return render(request, 'hr_management/invitation_view.html', context)
            
            # For API/JSON response (maintaining backward compatibility)
            invitation_data = {
                'invitation_id': str(invitation.invitation_id),
                'name': invitation.name,
                'email': invitation.email,
                'department': invitation.department.name if invitation.department else None,
                'designation': invitation.designation.name if invitation.designation else None,
                'role': invitation.role.name if invitation.role else None,
                'status': invitation.status,
                'created_at': invitation.created_at.isoformat() if invitation.created_at else None,
                'sent_at': invitation.sent_at.isoformat() if invitation.sent_at else None,
                'completed_at': invitation.completed_at.isoformat() if invitation.completed_at else None,
                'rejected_at': invitation.rejected_at.isoformat() if invitation.rejected_at else None,
                'accepted_at': invitation.accepted_at.isoformat() if invitation.accepted_at else None,
                'rejection_reason': invitation.rejection_reason,
                'discussion_message': invitation.discussion_message,
                'is_form_completed': invitation.is_form_completed,
                'has_viewed_offer': invitation.has_viewed_offer,
                'policies': invitation.policies,
            }
            
            # Include form data if the form is completed
            if invitation.is_form_completed:
                # Get form data from the policies field
                form_data = invitation.policies.get('form_data', {}) if invitation.policies else {}
                if form_data:
                    invitation_data['form_data'] = form_data
                
                # If no form data found in policies, try to get from associated employee
                elif not form_data:
                    try:
                        employee = Employee.objects.filter(employee_email=invitation.email).first()
                        if employee:
                            # Extract relevant employee information as form data
                            form_data = {
                                'full_name': employee.employee_name,
                                'email': employee.employee_email,
                                'phone_number': getattr(employee, 'phone_number', None),
                                'address': getattr(employee, 'address', None),
                                'date_of_birth': getattr(employee, 'date_of_birth', None).isoformat() if hasattr(employee, 'date_of_birth') and getattr(employee, 'date_of_birth') else None,
                                'emergency_contact': getattr(employee, 'emergency_contact', None),
                                'highest_education': getattr(employee, 'highest_education', None),
                                'previous_employer': getattr(employee, 'previous_employer', None),
                            }
                            invitation_data['form_data'] = form_data
                    except Exception as e:
                        logger.error(f"Error retrieving form data from employee: {str(e)}", exc_info=True)
            
            return JsonResponse({
                'success': True,
                'invitation': invitation_data
            })
            
        except OnboardingInvitation.DoesNotExist:
            if 'application/json' not in request.META.get('HTTP_ACCEPT', '') and not request.GET.get('format') == 'json' and not request.GET.get('v') == '2':
                # For HTML template
                context = {
                    'error': 'Invitation not found',
                    'current_year': timezone.now().year,
                }
                return render(request, 'hr_management/error.html', context)
            
            # For API/JSON
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error retrieving invitation details: {str(e)}", exc_info=True)
            
            if 'application/json' not in request.META.get('HTTP_ACCEPT', '') and not request.GET.get('format') == 'json' and not request.GET.get('v') == '2':
                # For HTML template
                context = {
                    'error': 'An error occurred',
                    'current_year': timezone.now().year,
                }
                return render(request, 'hr_management/error.html', context)
            
            # For API/JSON
            return JsonResponse({
                'success': False,
                'message': f'An error occurred while fetching the invitation details: {str(e)}'
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationAcceptView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        try:
            invitation = get_object_or_404(OnboardingInvitation, invitation_id=invitation_id)

            # Check if invitation is in a state to be accepted
            if invitation.status != 'completed':
                return JsonResponse({'success': False, 'message': f'This invitation cannot be accepted. Its status is "{invitation.status}".'}, status=400)

            # Extract data from the request
            gross_salary = request.POST.get('gross_salary')
            group_id = request.POST.get('group')
            payouts_str = request.POST.get('payouts')

            # Validate data
            if not gross_salary:
                return JsonResponse({'success': False, 'message': 'Gross salary is required.'}, status=400)

            # Create a password for the employee
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
            hashed_password = make_password(password)

            # Get leave group
            leave_group = None
            if group_id:
                try:
                    leave_group = LeaveGroup.objects.get(group_id=group_id)
                except (LeaveGroup.DoesNotExist, ValueError):
                    return JsonResponse({'success': False, 'message': 'Invalid Leave Group selected.'}, status=400)
            
            # Parse payouts
            payouts_data = json.loads(payouts_str) if payouts_str else []

            # Get the employee record that was created when the candidate submitted the form
            try:
                employee = Employee.objects.get(employee_email=invitation.email)
            except Employee.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Employee record not found. The candidate may not have completed the onboarding form.'}, status=404)

            # Update the employee record with the details from the acceptance modal
            employee.gross_salary = gross_salary
            employee.leave_group = leave_group
            employee.payouts = payouts_data
            employee.password = hashed_password
            employee.is_active = True
            employee.is_approved = True
            employee.save()

            # Update invitation status
            invitation.status = 'accepted'
            invitation.accepted_at = timezone.now()
            invitation.save()

            # Send acceptance email with credentials
            self.send_acceptance_email(invitation, password)
            
            return JsonResponse({'success': True, 'message': 'Invitation accepted and employee created successfully.'})

        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invitation not found.'}, status=404)
        except Exception as e:
            logger.error(f"Error accepting invitation: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'An unexpected error occurred: {e}'}, status=500)

    def send_acceptance_email(self, invitation, password):
        """Send email with login credentials to the accepted employee"""
        company = invitation.company
        
        # Get site URL from settings or use a default
        site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
        
        # Prepare context data for the email template
        context = {
            'name': invitation.name,
            'email': invitation.email,
            'company_name': company.company_name,
            'department': invitation.department.name if invitation.department else 'Not specified',
            'designation': invitation.designation.name if invitation.designation else 'Not specified',
            'role': invitation.role.name if invitation.role else 'Not specified',
            'login_url': f"{site_url}/hr/employee/login/",
            'username': invitation.email,  # Email is used as username
            'password': password,
            'current_year': timezone.now().year,
        }
        
        # Add company logo if available
        if company.company_logo:
            context['company_logo'] = company.company_logo.url
        
        try:
            # Render email template with context
            html_content = render_to_string('hr_management/email_templates/acceptance_credentials.html', context)
            text_content = strip_tags(html_content)
            
            # Send email
            subject = f"Welcome to {company.company_name} - Your Login Credentials"
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [invitation.email]
            
            send_mail(
                subject=subject,
                message=text_content,
                from_email=from_email,
                recipient_list=recipient_list,
                html_message=html_content,
                fail_silently=False,
            )
            
            logger.info(f"Acceptance email with credentials sent to {invitation.email}")
            return True
        except Exception as e:
            logger.error(f"Error sending acceptance email: {str(e)}", exc_info=True)
            # Don't raise the exception as the employee is already created
            # Just log the error for monitoring
            return False

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationRejectView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        """Reject an invitation"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Parse request data if present
            try:
                data = json.loads(request.body)
                action = data.get('action', '')
                from_status = data.get('from_status', '')
                to_status = data.get('to_status', '')
                rejection_reason = data.get('rejection_reason', '')
            except (ValueError, json.JSONDecodeError):
                data = {}
                action = ''
                from_status = ''
                to_status = ''
                rejection_reason = ''
            
            # Check if invitation can be rejected
            if invitation.status in ['accepted', 'cancelled']:
                return JsonResponse({
                    'success': False,
                    'message': f'This invitation cannot be rejected as it is {invitation.status}'
                }, status=400)

            # Save the rejection reason
            invitation.rejection_reason = rejection_reason
            
            # Update invitation status
            invitation.status = 'rejected'
            invitation.rejected_at = timezone.now()
            invitation.save()
            
            # Send rejection email
            email_sent = self.send_rejection_email(invitation)
            
            if email_sent:
                logger.info(f"Rejected invitation for {invitation.email} and sent notification")
            else:
                logger.warning(f"Rejected invitation for {invitation.email} but failed to send notification")
            
            return JsonResponse({
                'success': True,
                'message': 'Invitation rejected and notification sent',
                'email_sent': email_sent
            })
            
        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error rejecting invitation: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    def send_rejection_email(self, invitation):
        """Send rejection email to the applicant"""
        try:
            company = invitation.company
            
            # Get site URL from settings or use a default
            site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
            
            # Prepare context data for the email template
            context = {
                'name': invitation.name,
                'email': invitation.email,
                'company_name': company.company_name,
                'rejection_reason': invitation.rejection_reason or 'No specific reason provided',
                'current_year': timezone.now().year,
            }
            
            # Add company logo if available
            if company.company_logo:
                context['company_logo'] = company.company_logo.url
            
            # Check if template exists
            template_path = 'hr_management/email_templates/rejection_notification.html'
            
            # Render email template
            html_message = render_to_string(template_path, context)
            plain_message = strip_tags(html_message)
            
            # Send email
            subject = f"Update on Your Application at {company.company_name}"
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [invitation.email]
            
            email = EmailMultiAlternatives(subject, plain_message, from_email, recipient_list)
            email.attach_alternative(html_message, "text/html")
            email.send()
            
            logger.info(f"Rejection email sent to {invitation.email}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending rejection email: {str(e)}", exc_info=True)
            return False

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationDeleteView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        """Delete an invitation"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Save details for logging
            name = invitation.name
            email = invitation.email
            
            # Delete the invitation
            invitation.delete()
            
            # Log the deletion
            logger.info(f"Invitation for {name} ({email}) was deleted")
            
            return JsonResponse({
                'success': True,
                'message': 'Invitation deleted successfully'
            })
        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error deleting invitation: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
        
def get_employee_leave_balance(employee):
    leave_balance = {
        'casual': 0, 'sick': 0, 'annual': 0,
    }
    leave_policy = {
        'casual': 0, 'sick': 0, 'annual': 0,
    }
    
    if employee.leave_group:
        leave_group_rules = employee.leave_group.rules.all().select_related('leave_type')
        
        for rule in leave_group_rules:
            leave_type_name = rule.leave_type.name.lower()
            days_allowed = rule.days if rule.days is not None else 0
            if 'casual' in leave_type_name:
                leave_policy['casual'] = days_allowed
            elif 'sick' in leave_type_name:
                leave_policy['sick'] = days_allowed
            elif 'annual' in leave_type_name:
                leave_policy['annual'] = days_allowed
        
        current_year = timezone.now().year
        approved_leaves = LeaveApplication.objects.filter(
            employee=employee,
            status='approved',
            start_date__year=current_year
        ).select_related('leave_type')
        
        days_taken = {'casual': 0, 'sick': 0, 'annual': 0}
        
        for leave in approved_leaves:
            if leave.leave_type:
                leave_type_name = leave.leave_type.name.lower()
                duration = leave.duration if leave.duration is not None else 0
                if 'casual' in leave_type_name:
                    days_taken['casual'] += duration
                elif 'sick' in leave_type_name:
                    days_taken['sick'] += duration
                elif 'annual' in leave_type_name:
                    days_taken['annual'] += duration

        leave_balance['casual'] = leave_policy['casual'] - days_taken['casual']
        leave_balance['sick'] = leave_policy['sick'] - days_taken['sick']
        leave_balance['annual'] = leave_policy['annual'] - days_taken['annual']

    leave_balance['total'] = leave_balance['casual'] + leave_balance['sick'] + leave_balance['annual']
    
    return leave_balance, leave_policy

@class_employee_login_required
class EmployeeDashboardView(TemplateView):
    template_name = 'hr_management/employee_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get employee from session or request
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            # Redirect to login if no employee is logged in
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Today's date
            today = timezone.now().date()
            context['today_date'] = today
            
            # Attendance statistics
            month_start = today.replace(day=1)
            month_days = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            month_days = month_days.day
            
            # Get attendance for current month
            attendances = EmployeeAttendance.objects.filter(
                employee=employee,
                date__gte=month_start,
                date__lte=today
            )
            
            present_days = attendances.filter(status='present').count()
            
            context['attendance_stats'] = {
                'present': present_days,
                'total': month_days,
            }
            
            # Check today's status
            today_attendance = attendances.filter(date=today).first()
            context['today_status'] = today_attendance.status if today_attendance else 'not_marked'
            context['checked_out'] = today_attendance.check_out_time if today_attendance else False
            
            # Leave balance (example values, should be calculated from policy)
            context['leave_balance'] = {
                'casual': 7,
                'sick': 5,
                'annual': 14,
                'total': 26,
            }
            
            context['leave_policy'] = {
                'casual': 10,
                'sick': 7,
                'annual': 15,
            }
            
            # Pending requests
            pending_leaves = LeaveApplication.objects.filter(
                employee=employee, 
                status='pending'
            ).count()
            
            pending_reimbursements = ReimbursementRequest.objects.filter(
                employee=employee, 
                status='pending'
            ).count()
            
            context['pending_counts'] = {
                'leave': pending_leaves,
                'reimbursement': pending_reimbursements,
                'total': pending_leaves + pending_reimbursements,
            }
            
            # Latest salary slip
            latest_slip = SalarySlip.objects.filter(
                employee=employee
            ).order_by('-year', '-month').first()
            
            context['last_salary'] = latest_slip
            context['latest_slip'] = latest_slip
            
            # Recent Activities (collect recent activities for the employee)
            activities = []
            
            # Recent attendances
            recent_attendances = EmployeeAttendance.objects.filter(
                employee=employee
            ).order_by('-date', '-check_in_time')[:5]
            
            for attendance in recent_attendances:
                activity_type = "check_out" if attendance.check_out_time else "check_in"
                time_display = attendance.check_out_time if activity_type == "check_out" else attendance.check_in_time
                if time_display:
                    time_str = time_display.strftime('%I:%M %p')
                    activities.append({
                        'type': activity_type,
                        'icon': 'check' if activity_type == "check_in" else 'sign-out-alt',
                        'color': 'green' if activity_type == "check_in" else 'blue',
                        'title': 'Attendance Marked' if activity_type == "check_in" else 'Checked Out',
                        'description': f'You checked in at {time_str}' if activity_type == "check_in" else f'You checked out at {time_str}',
                        'timestamp': time_display,
                        'time_ago': self.get_time_ago(time_display)
                    })
            
            # Recent leave applications
            recent_leaves = LeaveApplication.objects.filter(
                employee=employee
            ).order_by('-created_at')[:3]
            
            for leave in recent_leaves:
                status_colors = {
                    'pending': 'amber',
                    'approved': 'green',
                    'rejected': 'red',
                    'cancelled': 'slate'
                }
                color = status_colors.get(leave.status, 'amber')
                activities.append({
                    'type': 'leave',
                    'icon': 'calendar-check',
                    'color': color,
                    'title': f'Leave Request {leave.status.capitalize()}',
                    'description': f'Your {leave.leave_type} leave from {leave.start_date.strftime("%d %b")} to {leave.end_date.strftime("%d %b")}',
                    'timestamp': leave.created_at,
                    'time_ago': self.get_time_ago(leave.created_at)
                })
            
            # Recent reimbursement requests
            recent_reimbursements = ReimbursementRequest.objects.filter(
                employee=employee
            ).order_by('-created_at')[:3]
            
            for reimbursement in recent_reimbursements:
                status_colors = {
                    'pending': 'amber',
                    'approved': 'green',
                    'rejected': 'red'
                }
                color = status_colors.get(reimbursement.status, 'amber')
                activities.append({
                    'type': 'reimbursement',
                    'icon': 'receipt',
                    'color': color,
                    'title': f'Reimbursement {reimbursement.status.capitalize()}',
                    'description': f'{reimbursement.currency} {reimbursement.amount} for {reimbursement.category}',
                    'timestamp': reimbursement.created_at,
                    'time_ago': self.get_time_ago(reimbursement.created_at)
                })
            
            # Recent salary slips
            if latest_slip:
                activities.append({
                    'type': 'salary',
                    'icon': 'file-invoice-dollar',
                    'color': 'emerald',
                    'title': 'Salary Slip Generated',
                    'description': f'{latest_slip.get_month_name} {latest_slip.year} salary slip has been generated',
                    'timestamp': latest_slip.created_at,
                    'time_ago': self.get_time_ago(latest_slip.created_at)
                })
            
            # Sort activities by timestamp (most recent first)
            activities.sort(key=lambda x: x['timestamp'], reverse=True)
            
            # Take only the top 10 activities
            context['activities'] = activities[:10]
            context['recent_activities'] = activities[:3]  # For the dashboard card
            
            # Get any pending invitations for this employee's email
            pending_invitations = OnboardingInvitation.objects.filter(
                email=employee.employee_email,
                status__in=['pending', 'sent']
            ).order_by('-created_at')
            
            context['invitations'] = pending_invitations
            
        except (Employee.DoesNotExist, Exception) as e:
            logger.error(f"Error preparing employee dashboard: {str(e)}", exc_info=True)
        
        return context
        
    def get_time_ago(self, timestamp):
        """Helper method to calculate human-readable time ago string"""
        if not timestamp:
            return ''
            
        now = timezone.now()
        diff = now - timestamp
        
        if diff.days > 365:
            years = diff.days // 365
            return f"{years}y ago"
        elif diff.days > 30:
            months = diff.days // 30
            return f"{months}mo ago"
        elif diff.days > 0:
            return f"{diff.days}d ago"
        elif diff.seconds > 3600:
            hours = diff.seconds // 3600
            return f"{hours}h ago"
        elif diff.seconds > 60:
            minutes = diff.seconds // 60
            return f"{minutes}m ago"
        else:
            return "just now"

@class_employee_login_required
class EmployeeProfileView(TemplateView):
    template_name = 'hr_management/employee_profile.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
        except Exception as e:
            logger.error(f"Error retrieving employee profile: {str(e)}", exc_info=True)
        
        return context
    
    def post(self, request, *args, **kwargs):
        # Handle profile update
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Update basic info
            if 'name' in request.POST:
                employee.name = request.POST.get('name')
            if 'phone' in request.POST:
                employee.phone_number = request.POST.get('phone')
            if 'address' in request.POST:
                employee.address = request.POST.get('address')
            
            # Handle profile photo if uploaded
            if 'photo' in request.FILES:
                employee.profile_photo = request.FILES.get('photo')
            
            employee.save()
            
            return JsonResponse({'success': True, 'message': 'Profile updated successfully'})
        except Exception as e:
            logger.error(f"Error updating employee profile: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@class_employee_login_required
class EmployeeLeaveView(TemplateView):
    template_name = 'hr_management/employee_leave.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Get all leave applications for this employee
            leave_applications = LeaveApplication.objects.filter(
                employee=employee
            ).order_by('-created_at')
            
            context['leave_applications'] = leave_applications
            
            # Get leave types for the company's user
            if employee.company and employee.company.user:
                context['leave_types'] = LeaveType.objects.filter(user=employee.company.user)
            else:
                context['leave_types'] = LeaveType.objects.none()
            
            # Leave balance (example values, should be calculated from policy)
            # Calculate actual leave balance from employee's leave group and used leaves
            leave_balance, leave_policy = get_employee_leave_balance(employee)
            context['leave_balance'] = leave_balance
            context['leave_policy'] = leave_policy
        except Exception as e:
            logger.error(f"Error retrieving employee leave data: {str(e)}", exc_info=True)
        
        return context

@class_employee_login_required
class EmployeeReimbursementView(TemplateView):
    template_name = 'hr_management/employee_reimbursement.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Get all reimbursement requests for this employee
            reimbursement_requests = ReimbursementRequest.objects.filter(
                employee=employee
            ).order_by('-created_at')
            
            context['reimbursement_requests'] = reimbursement_requests
            
            # Get expense categories
            context['expense_categories'] = dict(ReimbursementRequest.EXPENSE_CATEGORIES)
        except Exception as e:
            logger.error(f"Error retrieving employee reimbursement data: {str(e)}", exc_info=True)
        
        return context

@class_employee_login_required
class EmployeeSalarySlipsView(TemplateView):
    template_name = 'hr_management/employee_salary_slips.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Get all salary slips for this employee
            salary_slips = SalarySlip.objects.filter(
                employee=employee
            ).order_by('-year', '-month')
            
            context['salary_slips'] = salary_slips
        except Exception as e:
            logger.error(f"Error retrieving employee salary slips: {str(e)}", exc_info=True)
        
        return context

@class_employee_login_required
class EmployeeResignationView(TemplateView):
    template_name = 'hr_management/employee_resignation.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Check if employee has any existing resignation requests
            resignation = Resignation.objects.filter(
                employee=employee
            ).order_by('-created_at').first()
            
            context['resignation'] = resignation
            context['has_resignation'] = resignation is not None
            
            # Get resignation status choices
            context['resignation_statuses'] = dict(Resignation.RESIGNATION_STATUS)
        except Exception as e:
            logger.error(f"Error retrieving employee resignation data: {str(e)}", exc_info=True)
        
        return context

@class_employee_login_required
class EmployeeDocumentsView(TemplateView):
    template_name = 'hr_management/employee_documents.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.request.session.get('employee_id')
        
        if not employee_id:
            return context
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            context['employee'] = employee
            
            # Get all relevant documents
            context['company_policies'] = TandC.objects.all()
            context['handbooks'] = Handbook.objects.all()
            context['hiring_agreements'] = HiringAgreement.objects.all()
            context['training_materials'] = TrainingMaterial.objects.all()
        except Exception as e:
            logger.error(f"Error retrieving employee documents: {str(e)}", exc_info=True)
        
        return context

class EmployeeLogoutView(View):
    def post(self, request, *args, **kwargs):
        # Clear employee session
        if 'employee_id' in request.session:
            del request.session['employee_id']
        
        # Redirect to login page
        return redirect('/hr/employee/login/')

# API Views for Employee Dashboard
@method_decorator(csrf_exempt, name='dispatch')
class MarkAttendanceAPIView(View):
    def post(self, request, *args, **kwargs):
        # Get employee ID from session
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            # Try to parse JSON data first
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
            
            action = data.get('action', 'check_in')
            method = data.get('method', 'manual')
            location = data.get('location', {})
            qr_code = data.get('qr_code', {})
            
            # Ensure qr_code is a dictionary
            if isinstance(qr_code, str):
                try:
                    qr_code = json.loads(qr_code)
                except json.JSONDecodeError:
                    # If it's not valid JSON, create a basic structure
                    qr_code = {'raw_data': qr_code, 'qr_code_id': qr_code}
            
            # Get employee
            try:
                employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Employee not found'
                }, status=404)
            
            today = timezone.now().date()
            
            # Check if attendance already exists for today
            attendance = EmployeeAttendance.objects.filter(
                employee=employee,
                date=today
            ).first()
            
            # Validate check-out action
            if action == 'check_out' and (not attendance or not attendance.check_in_time):
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot check out without checking in first'
                }, status=400)
            
            # Validate check-in action
            if action == 'check_in' and attendance and attendance.check_in_time and not attendance.check_out_time:
                return JsonResponse({
                    'success': False,
                    'message': 'Already checked in for today, please check out first'
                }, status=400)
            
            # Validate if already checked out
            if attendance and attendance.check_out_time:
                return JsonResponse({
                    'success': False,
                    'message': 'Attendance already marked for today and checked out'
                }, status=400)
            
            # Extract location information if provided
            location_string = ""
            if isinstance(location, dict) and location.get('latitude') and location.get('longitude'):
                location_string = f"lat:{location.get('latitude')},lng:{location.get('longitude')}"
            
            # Handle QR code validation if method is 'qr'
            if method == 'qr' and qr_code:
                # Check for QR code ID - allow flexible formats
                qr_code_id = None
                if isinstance(qr_code, dict):
                    # Try different possible key names for QR code ID
                    for key in ['qr_code_id', 'id', 'code_id', 'raw_data']:
                        if key in qr_code and qr_code[key]:
                            qr_code_id = qr_code[key]
                            break
                
                # If no QR code ID found or QR code is not a dict, use a fallback
                if not qr_code_id:
                    if isinstance(qr_code, str):
                        qr_code_id = qr_code
                    else:
                        qr_code_id = str(uuid.uuid4())  # Generate a random ID as last resort
                
                # Try to validate QR code if possible
                try:
                    # Check if this QR code exists in the database
                    qr_code_obj = QRCode.objects.filter(qr_code_id=qr_code_id).first()
                    
                    # If QR code found, validate location
                    if qr_code_obj and isinstance(location, dict) and 'latitude' in location and 'longitude' in location:
                        # Get QR code location coordinates
                        qr_location = qr_code_obj.location_and_coordinates.get('coordinates', {})
                        qr_lat = qr_location.get('latitude')
                        qr_lng = qr_location.get('longitude')
                        
                        # Get user's submitted location
                        user_lat = location.get('latitude')
                        user_lng = location.get('longitude')
                        
                        # If both locations available, calculate distance
                        if qr_lat and qr_lng and user_lat and user_lng:
                            # Calculate distance in meters
                            distance = calculate_distance(qr_lat, qr_lng, user_lat, user_lng) * 1000
                            
                            # Check if user is within the allowed radius (20 meters)
                            # You can adjust this value based on your requirements
                            max_distance = 20  # meters
                            
                            if distance > max_distance:
                                return JsonResponse({
                                    'success': False, 
                                    'message': f'You are too far from the attendance location. Please be within {max_distance} meters to mark attendance.'
                                }, status=400)
                except Exception as e:
                    logger.warning(f"QR code validation error: {str(e)}")
                    # Continue with attendance marking even if validation fails
                    # In a production system, you might want to handle this differently
            
            # Create or update attendance record
            if attendance:
                # Update existing attendance record (check out)
                attendance.check_out_time = timezone.now()
                attendance.check_out_method = method
                attendance.check_out_location = location_string
                attendance.save()
                message = 'Checked out successfully'
            else:
                # Create new attendance record (check in)
                attendance = EmployeeAttendance.objects.create(
                    employee=employee,
                    date=today,
                    status='present',
                    check_in_time=timezone.now(),
                    check_in_method=method,
                    check_in_location=location_string,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    device_info=request.META.get('HTTP_USER_AGENT')
                )
                message = 'Attendance marked successfully'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'attendance_id': str(attendance.attendance_id),
                'check_in': attendance.check_in_time.isoformat() if attendance.check_in_time else None,
                'check_out': attendance.check_out_time.isoformat() if attendance.check_out_time else None,
                'action': 'check_out' if attendance.check_out_time else 'check_in'
            })
            
        except Exception as e:
            logger.error(f"Error marking attendance: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class LeaveApplicationAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            data = json.loads(request.body)
            leave_type_id = data.get('leave_type_id')
            start_date_str = data.get('start_date')
            end_date_str = data.get('end_date')
            reason = data.get('reason')
            
            # Validate required fields
            if not all([leave_type_id, start_date_str, end_date_str, reason]):
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            employee = Employee.objects.get(employee_id=employee_id)
            leave_type = LeaveType.objects.get(leave_type_id=leave_type_id, company=employee.company)

            # Parse dates
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format. Use YYYY-MM-DD'
                }, status=400)
            
            # Validate date range
            today = timezone.now().date()
            if start_date < today:
                return JsonResponse({
                    'success': False,
                    'message': 'Start date cannot be in the past'
                }, status=400)
            
            if end_date < start_date:
                return JsonResponse({
                    'success': False,
                    'message': 'End date cannot be before start date'
                }, status=400)
            
            # Calculate leave days
            leave_days = (end_date - start_date).days + 1
            
            
            # Check if there are overlapping leave applications
            overlapping_leaves = LeaveApplication.objects.filter(
                employee=employee,
                status__in=['pending', 'approved'],
                start_date__lte=end_date,
                end_date__gte=start_date
            ).exists()
            
            if overlapping_leaves:
                return JsonResponse({
                    'success': False,
                    'message': 'You already have a leave application for these dates'
                }, status=400)
            
            # Create leave application
            leave = LeaveApplication.objects.create(
                employee=employee,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                reason=reason,
                status='pending'
            )
            
            # Optional: Save attachments if provided
            attachments = request.FILES.getlist('attachments')
            if attachments:
                for attachment in attachments:
                    LeaveApplication.objects.create(
                        leave=leave,
                        file=attachment
                    )
            
            return JsonResponse({
                'success': True,
                'message': 'Leave application submitted successfully',
                'leave_id': str(leave.leave_id)
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Error submitting leave application: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class CancelLeaveAPIView(View):
    def post(self, request, leave_id, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            # Get employee
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Get the leave application
            try:
                leave = LeaveApplication.objects.get(leave_id=leave_id, employee=employee)
            except LeaveApplication.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Leave application not found'
                }, status=404)
            
            # Check if leave can be cancelled
            if leave.status not in ['pending', 'approved']:
                return JsonResponse({
                    'success': False,
                    'message': f'Cannot cancel leave in {leave.status} status'
                }, status=400)
            
            # Check if leave has already started
            today = timezone.now().date()
            if leave.start_date <= today:
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot cancel a leave that has already started or completed'
                }, status=400)
            
            # Update leave status to cancelled
            leave.status = 'cancelled'
            leave.updated_at = timezone.now()
            leave.save()
            
            # Optional: Parse and save cancellation reason if provided
            try:
                data = json.loads(request.body)
                cancellation_reason = data.get('reason')
                if cancellation_reason:
                    # You might need to add a field to the model for cancellation reason
                    # For now, we'll append it to the review notes
                    leave.review_notes = f"Cancelled by employee. Reason: {cancellation_reason}"
                    leave.save()
            except json.JSONDecodeError:
                pass  # No reason provided, that's fine
            
            return JsonResponse({
                'success': True,
                'message': 'Leave application cancelled successfully'
            })
            
        except Exception as e:
            logger.error(f"Error cancelling leave application: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class ReimbursementRequestAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            # Parse JSON data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                # Handle form-data for file uploads
                data = request.POST.dict()
            
            # Get form fields
            amount = data.get('amount')
            currency = data.get('currency', 'INR')
            category = data.get('category')
            expense_date_str = data.get('expense_date')
            description = data.get('description')
            
            # Validate required fields
            if not all([amount, category, expense_date_str, description]):
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            # Validate amount is numeric
            try:
                amount = float(amount)
                if amount <= 0:
                    raise ValueError("Amount must be positive")
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid amount'
                }, status=400)
            
            # Parse expense date
            try:
                expense_date = datetime.strptime(expense_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format. Use YYYY-MM-DD'
                }, status=400)
            
            # Validate expense date not in future
            today = timezone.now().date()
            if expense_date > today:
                return JsonResponse({
                    'success': False,
                    'message': 'Expense date cannot be in the future'
                }, status=400)
            
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Create reimbursement request
            reimbursement = ReimbursementRequest.objects.create(
                employee=employee,
                amount=amount,
                currency=currency,
                category=category,
                expense_date=expense_date,
                description=description,
                status='pending'
            )
            
            # Handle file uploads
            receipts = request.FILES.getlist('receipts')
            if receipts:
                for receipt in receipts:
                    ReimbursementRequest.objects.create(
                        reimbursement=reimbursement,
                        file=receipt
                    )
            
            return JsonResponse({
                'success': True,
                'message': 'Reimbursement request submitted successfully',
                'reimbursement_id': str(reimbursement.reimbursement_id)
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Error submitting reimbursement request: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class DeleteDocumentAPIView(View):
    def delete(self, request, document_id, *args, **kwargs):
        try:
            # Get employee from session
            employee_id = request.session.get('employee_id')
            if not employee_id:
                return JsonResponse({'success': False, 'message': 'Authentication required'}, status=401)
            
            # Get the document
            document = EmployeeDocument.objects.get(id=document_id, employee_id=employee_id)
            
            # Delete the file
            if document.file:
                if os.path.isfile(document.file.path):
                    os.remove(document.file.path)
            
            # Delete the document
            document.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Document deleted successfully'
            })
            
        except EmployeeDocument.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Document not found'}, status=404)
        except Exception as e:
            logger.error(f"Error deleting document: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

# Employee login required decorator
def employee_login_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return redirect('hr_employee_login')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Class-based view decorator

class HREmployeeLoginView(View):
    def get(self, request):
        # If already logged in, redirect to dashboard
        if 'employee_id' in request.session:
            return redirect('employee_dashboard')
        
        return render(request, 'hr_management/employee_login.html')
    
    def post(self, request):
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        if not email or not password:
            messages.error(request, 'Please provide both email and password')
            return render(request, 'hr_management/employee_login.html')
            
        try:
            # Try to get employee with matching email
            employee = Employee.objects.get(employee_email=email)
            
            # Verify password
            if check_password(password, employee.password):
                # Set session
                request.session['employee_id'] = str(employee.employee_id)
                
                # Redirect to dashboard
                return redirect('/hr/employee/dashboard/')
            else:
                messages.error(request, 'Invalid password')
        except Employee.DoesNotExist:
            messages.error(request, 'No account found with this email')
            
        return render(request, 'hr_management/employee_login.html')

@method_decorator(csrf_exempt, name='dispatch')
class OnboardingInvitationsListView(View):
    def get(self, request, *args, **kwargs):
        try:
            # Get all invitations, ordered by most recent first
            invitations = OnboardingInvitation.objects.all().order_by('-created_at')
            
            # Convert to a list of dictionaries for JSON response
            invitations_data = []
            for invitation in invitations:
                invitations_data.append({
                    'invitation_id': str(invitation.invitation_id),
                    'name': invitation.name,
                    'email': invitation.email,
                    'department': invitation.department.name if invitation.department else '',
                    'designation': invitation.designation.name if invitation.designation else '',
                    'role': invitation.role.name if invitation.role else '',
                    'status': invitation.status,
                    'sent_at': invitation.sent_at.isoformat() if invitation.sent_at else None,
                    'form_link': invitation.form_link
                })
            
            return JsonResponse({
                'success': True,
                'invitations': invitations_data
            })
        except Exception as e:
            logger.error(f"Error fetching invitations: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f"Error fetching invitations: {str(e)}"
            }, status=500)

class ViewOfferView(TemplateView):
    template_name = 'hr_management/view_offer.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        token = kwargs.get('token')
        if not token:
            logger.error("No token provided for view offer")
            return context
            
        try:
            # Find invitation by token in form_link
            invitations = OnboardingInvitation.objects.filter(form_link__contains=token)
            if not invitations.exists():
                logger.error(f"No invitation found with token {token}")
                return context
                
            invitation = invitations.first()
            context['invitation'] = invitation
            
            # Check if the form has already been completed
            if invitation.is_form_completed:
                context['form_completed'] = True
                context['form_data'] = {}
                return context
            
            # Mark the invitation as viewed
            invitation.has_viewed_offer = True
            invitation.save(update_fields=['has_viewed_offer'])
            
            # Process offer letter
            if invitation.offer_letter_template:
                offer_letter_content = invitation.offer_letter_template.content
                # Replace placeholders with actual values
                offer_letter_content = offer_letter_content.replace('[Employee Name]', invitation.name)
                try:
                    offer_letter_content = invitation.offer_letter_template.content
                    # Replace placeholders with actual values
                    offer_letter_content = offer_letter_content.replace('[Employee Name]', invitation.name)
                    offer_letter_content = offer_letter_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    offer_letter_content = offer_letter_content.replace('[Company Name]', invitation.company.company_name)
                    offer_letter_content = offer_letter_content.replace('[Start Date]', 'To be determined')
                    
                    context['offer_letter_content'] = offer_letter_content
                    logger.info(f"Successfully processed offer letter content")
                except Exception as e:
                    logger.error(f"Error processing offer letter content: {str(e)}")
                    context['offer_letter_content'] = '<p>Error processing offer letter content.</p>'
            else:
                context['offer_letter_content'] = '<p>No offer letter template was selected for this invitation.</p>'
                
            # Get hiring agreement content
            hiring_agreement_id = None
            additional_documents = invitation.additional_documents
            # Check in additional_documents first (new approach)
            if isinstance(additional_documents, list):
                for doc in additional_documents:
                    if doc.get('type') == 'hiring_agreement':
                        hiring_agreement_id = doc.get('id')
                        break
            
            # Fallback to old field
            if not hiring_agreement_id and invitation.hiring_agreement_template:
                try:
                    hiring_agreement_content = invitation.hiring_agreement_template.content
                    # Replace placeholders with actual values
                    hiring_agreement_content = hiring_agreement_content.replace('[Employee Name]', invitation.name)
                    hiring_agreement_content = hiring_agreement_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    hiring_agreement_content = hiring_agreement_content.replace('[Company Name]', invitation.company.company_name)
                    hiring_agreement_content = hiring_agreement_content.replace('[Start Date]', 'To be determined')
                    
                    context['hiring_agreement_content'] = hiring_agreement_content
                    logger.info(f"Successfully processed hiring agreement content")
                except Exception as e:
                    logger.error(f"Error processing hiring agreement content: {str(e)}")
                    context['hiring_agreement_content'] = '<p>Error processing hiring agreement content.</p>'
            # Use additional_documents approach
            elif hiring_agreement_id:
                try:
                    agreement = HiringAgreement.objects.get(agreement_id=hiring_agreement_id)
                    hiring_agreement_content = agreement.content
                    # Replace placeholders with actual values
                    hiring_agreement_content = hiring_agreement_content.replace('[Employee Name]', invitation.name)
                    hiring_agreement_content = hiring_agreement_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    hiring_agreement_content = hiring_agreement_content.replace('[Company Name]', invitation.company.company_name)
                    hiring_agreement_content = hiring_agreement_content.replace('[Start Date]', 'To be determined')
                    
                    context['hiring_agreement_content'] = hiring_agreement_content
                    logger.info(f"Successfully processed hiring agreement content from additional_documents")
                except Exception as e:
                    logger.error(f"Error processing hiring agreement content from additional_documents: {str(e)}")
                    context['hiring_agreement_content'] = '<p>Error processing hiring agreement content.</p>'
            else:
                context['hiring_agreement_content'] = '<p>No hiring agreement template was selected for this invitation.</p>'
                
            # Get handbook content
            handbook_id = None
            # Check in additional_documents first (new approach)
            if isinstance(additional_documents, list):
                for doc in additional_documents:
                    if doc.get('type') == 'handbook':
                        handbook_id = doc.get('id')
                        break
                        
            # Fallback to old field
            if not handbook_id and invitation.handbook_template:
                try:
                    handbook_content = invitation.handbook_template.content
                    # Replace placeholders with actual values
                    handbook_content = handbook_content.replace('[Employee Name]', invitation.name)
                    handbook_content = handbook_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    handbook_content = handbook_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['handbook_content'] = handbook_content
                    logger.info(f"Successfully processed handbook content")
                except Exception as e:
                    logger.error(f"Error processing handbook content: {str(e)}")
                    context['handbook_content'] = '<p>Error processing handbook content.</p>'
            # Use additional_documents approach
            elif handbook_id:
                try:
                    handbook = Handbook.objects.get(handbook_id=handbook_id)
                    handbook_content = handbook.content
                    # Replace placeholders with actual values
                    handbook_content = handbook_content.replace('[Employee Name]', invitation.name)
                    handbook_content = handbook_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    handbook_content = handbook_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['handbook_content'] = handbook_content
                    logger.info(f"Successfully processed handbook content from additional_documents")
                except Exception as e:
                    logger.error(f"Error processing handbook content from additional_documents: {str(e)}")
                    context['handbook_content'] = '<p>Error processing handbook content.</p>'
            else:
                context['handbook_content'] = '<p>No handbook template was selected for this invitation.</p>'
                
            # Get HR policies content
            hr_policies_id = None
            # Check in additional_documents first (new approach)
            if isinstance(additional_documents, list):
                for doc in additional_documents:
                    if doc.get('type') == 'hr_policies':
                        hr_policies_id = doc.get('id')
                        break
                        
            # Fallback to old field
            if not hr_policies_id and invitation.hr_policies_template:
                try:
                    hr_policies_content = invitation.hr_policies_template.content
                    # Replace placeholders with actual values
                    hr_policies_content = hr_policies_content.replace('[Employee Name]', invitation.name)
                    hr_policies_content = hr_policies_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    hr_policies_content = hr_policies_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['hr_policies_content'] = hr_policies_content
                    logger.info(f"Successfully processed HR policies content")
                except Exception as e:
                    logger.error(f"Error processing HR policies content: {str(e)}")
                    context['hr_policies_content'] = '<p>Error processing HR policies content.</p>'
            # Use additional_documents approach
            elif hr_policies_id:
                try:
                    hr_policy = TandC.objects.get(tandc_id=hr_policies_id)
                    hr_policies_content = hr_policy.content
                    # Replace placeholders with actual values
                    hr_policies_content = hr_policies_content.replace('[Employee Name]', invitation.name)
                    hr_policies_content = hr_policies_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    hr_policies_content = hr_policies_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['hr_policies_content'] = hr_policies_content
                    logger.info(f"Successfully processed HR policies content from additional_documents")
                except Exception as e:
                    logger.error(f"Error processing HR policies content from additional_documents: {str(e)}")
                    context['hr_policies_content'] = '<p>Error processing HR policies content.</p>'
            else:
                context['hr_policies_content'] = '<p>No HR policies template was selected for this invitation.</p>'
                
            # Get training material content
            training_material_id = None
            # Check in additional_documents first (new approach)
            if isinstance(additional_documents, list):
                for doc in additional_documents:
                    if doc.get('type') == 'training_material':
                        training_material_id = doc.get('id')
                        break
                        
            # Fallback to old field
            if not training_material_id and invitation.training_material_template:
                try:
                    training_material_content = invitation.training_material_template.content
                    # Replace placeholders with actual values
                    training_material_content = training_material_content.replace('[Employee Name]', invitation.name)
                    training_material_content = training_material_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    training_material_content = training_material_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['training_material_content'] = training_material_content
                    logger.info(f"Successfully processed training material content")
                except Exception as e:
                    logger.error(f"Error processing training material content: {str(e)}")
                    context['training_material_content'] = '<p>Error processing training material content.</p>'
            # Use additional_documents approach
            elif training_material_id:
                try:
                    training_material = TrainingMaterial.objects.get(material_id=training_material_id)
                    training_material_content = training_material.content
                    # Replace placeholders with actual values
                    training_material_content = training_material_content.replace('[Employee Name]', invitation.name)
                    training_material_content = training_material_content.replace('[Designation]', invitation.designation.name if invitation.designation else '')
                    training_material_content = training_material_content.replace('[Company Name]', invitation.company.company_name)
                    
                    context['training_material_content'] = training_material_content
                    logger.info(f"Successfully processed training material content from additional_documents")
                except Exception as e:
                    logger.error(f"Error processing training material content from additional_documents: {str(e)}")
                    context['training_material_content'] = '<p>Error processing training material content.</p>'
            else:
                context['training_material_content'] = '<p>No training material template was selected for this invitation.</p>'
                
            # Get policies if any
            if invitation.policies:
                try:
                    # Handle various formats of the policies field
                    if isinstance(invitation.policies, str):
                        policy_ids = json.loads(invitation.policies)
                    elif isinstance(invitation.policies, list):
                        policy_ids = invitation.policies
                    elif isinstance(invitation.policies, dict):
                        # If it's a dict with policy IDs as keys or values
                        policy_ids = list(invitation.policies.keys()) if len(invitation.policies) > 0 else []
                    else:
                        policy_ids = []
                        logger.warning(f"Unexpected policies format: {type(invitation.policies)}")
                    
                    policies = []
                    for policy_id in policy_ids:
                        try:
                            policy = TandC.objects.get(tandc_id=policy_id)
                            policies.append(policy)
                        except TandC.DoesNotExist:
                            logger.warning(f"Policy with ID {policy_id} not found")
                            continue
                    context['policies'] = policies
                except json.JSONDecodeError:
                    logger.error(f"Invalid JSON in policies field: {invitation.policies}")
                except Exception as e:
                    logger.error(f"Error processing policies: {str(e)}", exc_info=True)
                
        except Exception as e:
            logger.error(f"Error retrieving offer details: {str(e)}", exc_info=True)
            
        return context


@method_decorator(csrf_exempt, name='dispatch')
class OfferResponseView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        try:
            # Get the invitation
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Check if the form has already been completed
            if invitation.is_form_completed:
                return JsonResponse({
                    'success': False,
                    'message': 'This form has already been submitted and is no longer accessible.'
                }, status=403)
            
            # Get response type
            response = request.POST.get('response')
            
            if response == 'accept':
                # Update status to completed
                invitation.status = 'completed'  # Keep as 'completed' for backward compatibility
                invitation.completed_at = timezone.now()
                invitation.save()
                
                # Redirect to onboarding form
                token = invitation.form_link.split('/')[-2]
                return redirect(f"/hr/onboarding/form/{token}/")
                
            elif response == 'reject':
                # Get rejection reason
                reason = request.POST.get('reason', '')
                
                # Update status to rejected
                invitation.status = 'rejected'
                invitation.rejected_at = timezone.now()
                invitation.rejection_reason = reason
                invitation.save()
                
                # Send notification email to HR
                try:
                    self.send_rejection_notification(invitation)
                except Exception as e:
                    logger.error(f"Failed to send rejection notification: {str(e)}", exc_info=True)
                
                # Return a success page
                return render(request, 'hr_management/offer_response_success.html', {
                    'response': 'reject',
                    'company_name': invitation.company.company_name,
                    'invitation_id': invitation.invitation_id
                })
                
            elif response == 'discuss':
                # Get discussion message
                message = request.POST.get('message', '')
                
                # Update status to need discussion
                invitation.status = 'need_discussion'
                invitation.discussion_message = message
                invitation.save()
                
                # Send notification email to HR
                try:
                    self.send_discussion_notification(invitation)
                except Exception as e:
                    logger.error(f"Failed to send discussion notification: {str(e)}", exc_info=True)
                
                # Return a success page
                return render(request, 'hr_management/offer_response_success.html', {
                    'response': 'discuss',
                    'company_name': invitation.company.company_name,
                    'invitation_id': invitation.invitation_id
                })
                
            else:
                logger.error(f"Invalid response type: {response}")
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid response type'
                }, status=400)
                
        except OnboardingInvitation.DoesNotExist:
            logger.error(f"Invitation not found: {invitation_id}")
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
            
        except Exception as e:
            logger.error(f"Error processing offer response: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    def send_rejection_notification(self, invitation):
        """Send notification email to HR when offer is rejected"""
        try:
            company = invitation.company
            hr_email = company.company_email
            
            if not hr_email:
                logger.warning(f"No HR email found for company {company.company_name}")
                return False
            
            # Get site URL from settings or use a default
            site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
            dashboard_url = f"{site_url}/hr/onboarding/"
            
            # Prepare context for the email template
            context = {
                'name': invitation.name,
                'email': invitation.email,
                'department': invitation.department.name if invitation.department else 'Not specified',
                'designation': invitation.designation.name if invitation.designation else 'Not specified',
                'role': invitation.role.name if invitation.role else 'Not specified',
                'rejection_reason': invitation.rejection_reason or 'No reason provided',
                'dashboard_url': dashboard_url,
                'current_year': timezone.now().year,
            }
            
            # Render email templates
            html_content = render_to_string('hr_management/email_templates/rejection_notification.html', context)
            text_content = strip_tags(html_content)
            
            # Create email message
            subject = f"Offer Rejected by {invitation.name}"
            from_email = settings.DEFAULT_FROM_EMAIL
            
            msg = EmailMultiAlternatives(subject, text_content, from_email, [hr_email])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            
            logger.info(f"Rejection notification sent to HR for {invitation.name}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending rejection notification: {str(e)}", exc_info=True)
            raise
    
    def send_discussion_notification(self, invitation):
        """Send notification email to HR when candidate requests discussion"""
        try:
            company = invitation.company
            hr_email = company.company_email
            
            if not hr_email:
                logger.warning(f"No HR email found for company {company.company_name}")
                return False
            
            # Get site URL from settings or use a default
            site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
            dashboard_url = f"{site_url}/hr/onboarding/"
            
            # Prepare context for the email template
            context = {
                'name': invitation.name,
                'email': invitation.email,
                'department': invitation.department.name if invitation.department else 'Not specified',
                'designation': invitation.designation.name if invitation.designation else 'Not specified',
                'role': invitation.role.name if invitation.role else 'Not specified',
                'discussion_message': invitation.discussion_message or 'No specific message provided',
                'dashboard_url': dashboard_url,
                'current_year': timezone.now().year,
            }
            
            # Render email templates
            html_content = render_to_string('hr_management/email_templates/discussion_notification.html', context)
            text_content = strip_tags(html_content)
            
            # Create email message
            subject = f"Discussion Requested by {invitation.name}"
            from_email = settings.DEFAULT_FROM_EMAIL
            
            msg = EmailMultiAlternatives(subject, text_content, from_email, [hr_email])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            
            logger.info(f"Discussion notification sent to HR for {invitation.name}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending discussion notification: {str(e)}", exc_info=True)
            raise

@method_decorator(csrf_exempt, name='dispatch')
class ResignationAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            # Parse the JSON-encoded request body to extract the data sent by the client
            data = json.loads(request.body)
            last_working_date = data.get('last_working_date')
            reason = data.get('reason')
            additional_notes = data.get('additional_notes')
            
            # Validate input
            if not all([last_working_date, reason]):
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            # Convert date
            try:
                last_working_date = datetime.strptime(last_working_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format. Use YYYY-MM-DD'
                }, status=400)
            
            # Validate notice period (at least 2 weeks)
            resignation_date = timezone.now().date()
            notice_period = (last_working_date - resignation_date).days
            
            if notice_period < 14:
                return JsonResponse({
                    'success': False,
                    'message': 'Notice period must be at least 14 days'
                }, status=400)
            
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Check if employee already has a pending resignation
            existing_resignation = Resignation.objects.filter(
                employee=employee,
                status__in=['pending', 'acknowledged', 'processing']
            ).exists()
            
            if existing_resignation:
                return JsonResponse({
                    'success': False,
                    'message': 'You already have a pending resignation request'
                }, status=400)
            
            # Create resignation request
            resignation = Resignation.objects.create(
                employee=employee,
                resignation_date=resignation_date,
                last_working_date=last_working_date,
                reason=reason,
                additional_notes=additional_notes,
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Resignation submitted successfully',
                'resignation_id': str(resignation.resignation_id),
                'status': resignation.status,
                'notice_period_days': resignation.notice_period_days
            })
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Error submitting resignation: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class CancelResignationAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            # Get employee
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Get the resignation request
            resignation = Resignation.objects.filter(
                employee=employee,
                status__in=['pending', 'acknowledged', 'processing']
            ).order_by('-created_at').first()
            
            if not resignation:
                return JsonResponse({
                    'success': False,
                    'message': 'No active resignation request found'
                }, status=404)
            
            # Get optional cancellation reason
            cancellation_reason = None
            try:
                data = json.loads(request.body)
                cancellation_reason = data.get('reason')
            except json.JSONDecodeError:
                pass  # No JSON data, that's fine
            
            # Update resignation status to cancelled
            resignation.status = 'cancelled'
            
            # Add cancellation reason to feedback if provided
            if cancellation_reason:
                resignation.feedback = f"Cancelled by employee. Reason: {cancellation_reason}"
            
            resignation.updated_at = timezone.now()
            resignation.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Resignation request cancelled successfully'
            })
            
        except Exception as e:
            logger.error(f"Error cancelling resignation: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class UpdateProfileAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Parse data based on content type
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                # Handle form-data for file uploads
                data = request.POST.dict()
                
            # Update basic information
            if 'name' in data:
                employee.name = data.get('name')
            if 'phone_number' in data:
                employee.phone_number = data.get('phone_number')
            if 'address' in data:
                employee.address = data.get('address')
            if 'emergency_contact_name' in data:
                employee.emergency_contact_name = data.get('emergency_contact_name')
            if 'emergency_contact_phone' in data:
                employee.emergency_contact_phone = data.get('emergency_contact_phone')
            if 'date_of_birth' in data:
                try:
                    date_of_birth = datetime.strptime(data.get('date_of_birth'), '%Y-%m-%d').date()
                    employee.date_of_birth = date_of_birth
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid date format for date of birth. Use YYYY-MM-DD'
                    }, status=400)
                    
            # Handle file uploads
            if 'photo' in request.FILES:
                employee.photo = request.FILES.get('photo')
                
            # Save changes
            employee.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Profile updated successfully'
            })
            
        except Exception as e:
            logger.error(f"Error updating profile: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class UpdatePasswordAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            data = json.loads(request.body)
            current_password = data.get('current_password')
            new_password = data.get('new_password')
            confirm_password = data.get('confirm_password')
            
            # Validate required fields
            if not all([current_password, new_password, confirm_password]):
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
                
            # Validate new password matches confirmation
            if new_password != confirm_password:
                return JsonResponse({
                    'success': False,
                    'message': 'New password and confirmation do not match'
                }, status=400)
                
            # Validate password complexity
            if len(new_password) < 8:
                return JsonResponse({
                    'success': False,
                    'message': 'Password must be at least 8 characters long'
                }, status=400)
                
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Verify current password
            from django.contrib.auth.hashers import check_password, make_password
            if not check_password(current_password, employee.password):
                return JsonResponse({
                    'success': False,
                    'message': 'Current password is incorrect'
                }, status=400)
                
            # Update password
            employee.password = make_password(new_password)
            employee.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Password updated successfully'
            })
            
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Error updating password: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class UploadDocumentAPIView(View):
    def post(self, request, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Get form data
            document_type = request.POST.get('document_type')
            document_name = request.POST.get('document_name')
            notes = request.POST.get('notes', '')
            
            # Validate required fields
            if not all([document_type, document_name]) or 'file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            document_file = request.FILES.get('file')
            
            # Validate file size (max 10MB)
            if document_file.size > 10 * 1024 * 1024:
                return JsonResponse({
                    'success': False,
                    'message': 'File size exceeds 10MB limit'
                }, status=400)
            
            # Create document
            document = EmployeeDocument.objects.create(
                employee=employee,
                document_type=document_type,
                document_name=document_name,
                notes=notes,
                file=document_file,
                file_size=f"{document_file.size / 1024:.1f} KB"
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Document uploaded successfully',
                'document_id': document.id,
                'document_name': document.document_name,
                'document_type': document.document_type,
                'upload_date': document.uploaded_at.strftime('%Y-%m-%d %H:%M')
            })
            
        except Exception as e:
            logger.error(f"Error uploading document: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class FormCompletedActionView(View):
    """
    View to handle actions when an employee completes an onboarding form.
    Provides the UI and API for accept/reject actions.
    """
    
    def get(self, request, invitation_id, *args, **kwargs):
        """Render the UI for Accept/Reject buttons after form completion"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Check if the form is completed
            if not invitation.is_form_completed:
                return JsonResponse({
                    'success': False,
                    'message': 'The form has not been completed yet'
                }, status=400)
                
            # If the invitation is already processed
            if invitation.status in ['accepted', 'rejected']:
                status_message = f"This application has already been {invitation.status}"
                return render(request, 'hr_management/form_action_result.html', {
                    'success': True,
                    'invitation': invitation,
                    'status_message': status_message
                })
            
            # Extract form data for display
            form_data = {}
            if invitation.policies and 'form_data' in invitation.policies:
                form_data = invitation.policies['form_data']
                
                # Extract structured data if available
                if 'structured_data' in form_data:
                    form_data['structured_data'] = form_data['structured_data']
                
                # Extract personal info if available
                if 'personal_info' in form_data:
                    form_data['personal_info'] = form_data['personal_info']
                
                # Extract employment details if available
                if 'employment_details' in form_data:
                    form_data['employment_details'] = form_data['employment_details']
                
                logger.info(f"Retrieved form data for invitation {invitation_id}")
            else:
                logger.warning(f"No form data found for completed invitation {invitation_id}")
            
            # Get leave types, deductions and allowances for this company
            company = invitation.company
            leave_types = LeaveType.objects.filter(company=company, is_active=True)
            deductions = Deduction.objects.filter(company=company, is_active=True)
            allowances = Allowance.objects.filter(company=company, is_active=True)
                
            # Render the UI for accept/reject options
            return render(request, 'hr_management/form_action.html', {
                'invitation': invitation,
                'form_data': form_data,
                'leave_types': leave_types,
                'deductions': deductions,
                'allowances': allowances
            })
            
        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error rendering form action UI: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    def post(self, request, invitation_id, *args, **kwargs):
        """Handle accept/reject actions"""
        try:
            invitation = OnboardingInvitation.objects.get(invitation_id=invitation_id)
            
            # Parse request data
            try:
                data = json.loads(request.body)
                action = data.get('action', '')
                rejection_reason = data.get('rejection_reason', '')
                # Get the employment configuration parameters
                salary_ctc = data.get('salary_ctc', 0)
                leave_types = data.get('leave_types', [])
                deductions = data.get('deductions', [])
                allowances = data.get('allowances', [])
            except (ValueError, json.JSONDecodeError):
                # Try to get from POST data if not JSON
                action = request.POST.get('action', '')
                rejection_reason = request.POST.get('rejection_reason', '')
                salary_ctc = request.POST.get('salary_ctc', 0)
                leave_types = request.POST.getlist('leave_types', [])
                deductions = request.POST.getlist('deductions', [])
                allowances = request.POST.getlist('allowances', [])
            
            # Check if the form is completed
            if not invitation.is_form_completed:
                return JsonResponse({
                    'success': False,
                    'message': 'The form has not been completed yet'
                }, status=400)
                
            # Check if already processed
            if invitation.status in ['accepted', 'rejected']:
                return JsonResponse({
                    'success': False,
                    'message': f'This application has already been {invitation.status}'
                }, status=400)
            
            if action == 'accept':
                # Validate required employment details
                if not salary_ctc:
                    return JsonResponse({
                        'success': False,
                        'message': 'Salary CTC is required for accepting the application'
                    }, status=400)
                
                # Use the model method for acceptance
                invitation.accept_invitation()
                
                # Generate random password
                password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                hashed_password = make_password(password)
                
                # Extract form data from the invitation's policies field
                form_data = {}
                if isinstance(invitation.policies, dict) and 'form_data' in invitation.policies:
                    form_data = invitation.policies['form_data']
                
                # Create or update employee record
                # Using hr.models.Employee (not employee.models.Employee)
                employee, created = Employee.objects.get_or_create(
                    employee_email=invitation.email,
                    defaults={
                        'employee_name': invitation.name,
                        'company': invitation.company,
                        # 'department': invitation.department,  # hr.models.Employee has no department field
                        # 'designation': invitation.designation,  # hr.models.Employee has no designation field
                        # 'role': invitation.role,  # hr.models.Employee has no role field
                        'is_active': True,
                        'is_approved': True,
                        'password': hashed_password
                    }
                )
                
                if not created:
                    # Update existing employee
                    employee.employee_name = invitation.name
                    employee.company = invitation.company
                    # employee.department = invitation.department  # hr.models.Employee has no department field
                    # employee.designation = invitation.designation  # hr.models.Employee has no designation field
                    # employee.role = invitation.role  # hr.models.Employee has no role field
                    employee.is_active = True
                    employee.is_approved = True
                    employee.password = hashed_password
                    employee.save()
                
                # Update salary and allowances in the employee record
                employee.salary_ctc = salary_ctc
                employee.allowances = allowances
                employee.save()
                
                # If there's a profile photo saved in the invitation, copy it to the employee
                if invitation.photo:
                    employee.attendance_photo = invitation.photo
                    employee.save()
                
                # Process any pending documents that were saved during form submission
                if isinstance(invitation.policies, dict) and 'document_files' in invitation.policies:
                    document_files = invitation.policies['document_files']
                    
                    # Create EmployeeDocument records for each stored document
                    for key, doc_info in document_files.items():
                        try:
                            # Create document with stored information
                            document_type = doc_info.get('document_type', '')
                            file_path = doc_info.get('file_path', '')
                            file_name = doc_info.get('file_name', '')
                            file_size = doc_info.get('file_size', '')
                            
                            if file_path and file_name:
                                # Create the document record linked to the employee
                                doc = EmployeeDocument(
                                    employee=employee,
                                    document_type=document_type,
                                    document_name=file_name,
                                    file=file_path,
                                    file_size=file_size,
                                )
                                doc.save()
                                logger.info(f"Created document record for {document_type} from stored file info")
                        except Exception as doc_error:
                            logger.error(f"Error creating document from stored info: {str(doc_error)}", exc_info=True)
                
                # Save employment configuration
                try:
                    # Add to policies JSON field
                    if not isinstance(invitation.policies, dict):
                        invitation.policies = {}
                    
                    # Add salary and benefits information
                    employment_config = {
                        'salary_ctc': float(salary_ctc),
                        'leave_types': leave_types,
                        'deductions': deductions,
                        'allowances': allowances,
                    }
                    
                    invitation.policies['employment_config'] = employment_config
                    invitation.save()
                    
                    # Create a salary record if SalarySlip model is used
                    try:
                        # Create a salary record
                        current_date = timezone.now()
                        
                        # Process allowances
                        allowance_dict = {}
                        total_allowances = 0
                        
                        for allowance_data in allowances:
                            allowance_id = allowance_data.get('id')
                            value = float(allowance_data.get('value', 0))
                            
                            if allowance_id and value > 0:
                                try:
                                    allowance = Allowance.objects.get(allowance_id=allowance_id)
                                    allowance_dict[allowance.name] = value
                                    total_allowances += value
                                except Allowance.DoesNotExist:
                                    continue
                        
                        # Process deductions
                        deduction_dict = {}
                        total_deductions = 0
                        
                        for deduction_data in deductions:
                            deduction_id = deduction_data.get('id')
                            value = float(deduction_data.get('value', 0))
                            
                            if deduction_id and value > 0:
                                try:
                                    deduction = Deduction.objects.get(deduction_id=deduction_id)
                                    deduction_dict[deduction.name] = value
                                    total_deductions += value
                                except Deduction.DoesNotExist:
                                    continue
                        
                        # Calculate net salary
                        basic_salary = float(salary_ctc)
                        net_salary = basic_salary + total_allowances - total_deductions
                        
                        # Create or update salary slip
                        salary_slip, created = SalarySlip.objects.get_or_create(
                            employee=employee,
                            month=current_date.month,
                            year=current_date.year,
                            defaults={
                                'basic_salary': basic_salary,
                                'allowances': allowance_dict,
                                'deductions': deduction_dict,
                                'net_salary': net_salary,
                                'payment_date': current_date,
                                'payment_method': 'Bank Transfer',
                                'is_paid': False,
                                'notes': 'Initial salary configuration',
                            }
                        )
                        
                        if not created:
                            # Update existing salary slip
                            salary_slip.basic_salary = basic_salary
                            salary_slip.allowances = allowance_dict
                            salary_slip.deductions = deduction_dict
                            salary_slip.net_salary = net_salary
                            salary_slip.save()
                        
                        logger.info(f"Created/updated salary configuration for employee {employee.employee_id}")
                    except Exception as salary_error:
                        logger.error(f"Error creating salary record: {str(salary_error)}", exc_info=True)
                except Exception as config_error:
                    logger.error(f"Error saving employment configuration: {str(config_error)}", exc_info=True)
                
                # Send acceptance email with login credentials
                success = self.send_acceptance_email(invitation, password)
                
                return JsonResponse({
                    'success': True,
                    'message': 'Invitation accepted successfully',
                    'email_sent': success
                })
                
            elif action == 'reject':
                # Update the invitation status
                invitation.status = 'rejected'
                invitation.rejected_at = timezone.now()
                invitation.rejection_reason = rejection_reason
                invitation.save()
            
                # Send rejection email
                success = self.send_rejection_email(invitation)
            
                return JsonResponse({
                    'success': True,
                    'message': 'Invitation rejected successfully',
                    'email_sent': success
                })
                
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'Invalid action: {action}'
                }, status=400)
            
        except OnboardingInvitation.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Invitation not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error processing form action: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    def send_acceptance_email(self, invitation, password):
        """Send email with login credentials to the accepted employee"""
        company = invitation.company
        
        # Get site URL from settings or use a default
        site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
        
        # Prepare context data for the email template
        context = {
            'name': invitation.name,
            'email': invitation.email,
            'company_name': company.company_name,
            'department': invitation.department.name if invitation.department else 'Not specified',
            'designation': invitation.designation.name if invitation.designation else 'Not specified',
            'role': invitation.role.name if invitation.role else 'Not specified',
            'login_url': f"{site_url}/hr/employee/login/",
            'username': invitation.email,  # Email is used as username
            'password': password,
            'current_year': timezone.now().year,
        }
        
        # Add company logo if available
        if company.company_logo:
            context['company_logo'] = company.company_logo.url
        
        try:
            # Render email template with context
            html_content = render_to_string('hr_management/email_templates/acceptance_credentials.html', context)
            text_content = strip_tags(html_content)
            
            # Send email
            subject = f"Welcome to {company.company_name} - Your Login Credentials"
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [invitation.email]
            
            send_mail(
                subject=subject,
                message=text_content,
                from_email=from_email,
                recipient_list=recipient_list,
                html_message=html_content,
                fail_silently=False,
            )
            
            logger.info(f"Acceptance email with credentials sent to {invitation.email}")
            return True
        except Exception as e:
            logger.error(f"Error sending acceptance email: {str(e)}", exc_info=True)
            # Don't raise the exception as the employee is already created
            # Just log the error for monitoring
            return False
    
    def send_rejection_email(self, invitation):
        """Send rejection email to the applicant"""
        try:
            company = invitation.company
            
            # Get site URL from settings or use a default
            site_url = getattr(settings, 'SITE_URL', "https://1matrix.io")
            
            # Prepare context data for the email template
            context = {
                'name': invitation.name,
                'email': invitation.email,
                'company_name': company.company_name,
                'rejection_reason': invitation.rejection_reason or 'No specific reason provided',
                'current_year': timezone.now().year,
            }
            
            # Add company logo if available
            if company.company_logo:
                context['company_logo'] = company.company_logo.url
            
            # Check if template exists
            template_path = 'hr_management/email_templates/rejection_notification.html'
            
            # Render email template
            html_message = render_to_string(template_path, context)
            plain_message = strip_tags(html_message)
            
            # Send email
            subject = f"Update on Your Application at {company.company_name}"
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [invitation.email]
            
            email = EmailMultiAlternatives(subject, plain_message, from_email, recipient_list)
            email.attach_alternative(html_message, "text/html")
            
            # Send the email
            email.send()
            
            logger.info(f"Rejection email sent to {invitation.email}")
            return True
            
        except Exception as e:
            logger.error(f"Error sending rejection email: {str(e)}", exc_info=True)
            return False

# Preview view for Training Material
class TrainingMaterialPreviewView(View):
    def get(self, request, template_id):
        try:
            template = TrainingMaterial.objects.get(training_material_id=template_id)
            return JsonResponse({
                'success': True,
                'content': template.content
            })
        except TrainingMaterial.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Update view for Training Material
class TrainingMaterialUpdateView(View):
    def post(self, request, template_id):
        try:
            template = TrainingMaterial.objects.get(training_material_id=template_id)
            name = request.POST.get('material_name')
            content = request.POST.get('material_content')
            
            if name and content:
                template.name = name
                template.content = content
                template.updated_at = timezone.now()
                template.save()
                return JsonResponse({'success': True})
            
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except TrainingMaterial.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Delete view for Training Material
class TrainingMaterialDeleteView(View):
    def post(self, request, template_id):
        try:
            template = TrainingMaterial.objects.get(training_material_id=template_id)
            template.delete()
            return JsonResponse({'success': True})
        except TrainingMaterial.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Update view for OfferLetter
class OfferLetterUpdateView(View):
    def post(self, request, template_id):
        try:
            template = OfferLetter.objects.get(offer_letter_id=template_id)
            name = request.POST.get('offer_letter_name')
            content = request.POST.get('offer_letter_content')
            
            if name and content:
                template.name = name
                template.content = content
                template.updated_at = timezone.now()
                template.save()
                return JsonResponse({'success': True})
            
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except OfferLetter.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Delete view for OfferLetter
class OfferLetterDeleteView(View):
    def post(self, request, template_id):
        try:
            template = OfferLetter.objects.get(offer_letter_id=template_id)
            template.delete()
            return JsonResponse({'success': True})
        except OfferLetter.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Preview view for Policy
class PolicyPreviewView(View):
    def get(self, request, template_id):
        try:
            template = TandC.objects.get(tandc_id=template_id)
            return JsonResponse({
                'success': True,
                'content': template.description
            })
        except TandC.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Update view for Policy
class PolicyUpdateView(View):
    def post(self, request, template_id):
        try:
            template = TandC.objects.get(tandc_id=template_id)
            name = request.POST.get('policy_name')
            description = request.POST.get('policy_description')
            
            if name and description:
                template.name = name
                template.description = description
                template.updated_at = timezone.now()
                template.save()
                return JsonResponse({'success': True})
            
            return JsonResponse({'error': 'Name and description are required'}, status=400)
        except TandC.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Delete view for Policy
class PolicyDeleteView(View):
    def post(self, request, template_id):
        try:
            template = TandC.objects.get(tandc_id=template_id)
            template.delete()
            return JsonResponse({'success': True})
        except TandC.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Update view for HiringAgreement
class HiringAgreementUpdateView(View):
    def post(self, request, template_id):
        try:
            template = HiringAgreement.objects.get(hiring_agreement_id=template_id)
            name = request.POST.get('agreement_name')
            content = request.POST.get('agreement_content')
            
            if name and content:
                template.name = name
                template.content = content
                template.updated_at = timezone.now()
                template.save()
                return JsonResponse({'success': True})
            
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except HiringAgreement.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Delete view for HiringAgreement
class HiringAgreementDeleteView(View):
    def post(self, request, template_id):
        try:
            template = HiringAgreement.objects.get(hiring_agreement_id=template_id)
            template.delete()
            return JsonResponse({'success': True})
        except HiringAgreement.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Update view for Handbook
class HandbookUpdateView(View):
    def post(self, request, template_id):
        try:
            template = Handbook.objects.get(handbook_id=template_id)
            name = request.POST.get('handbook_name')
            content = request.POST.get('handbook_content')
            
            if name and content:
                template.name = name
                template.content = content
                template.updated_at = timezone.now()
                template.save()
                return JsonResponse({'success': True})
            
            return JsonResponse({'error': 'Name and content are required'}, status=400)
        except Handbook.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

# Delete view for Handbook
class HandbookDeleteView(View):
    def post(self, request, template_id):
        try:
            template = Handbook.objects.get(handbook_id=template_id)
            template.delete()
            return JsonResponse({'success': True})
        except Handbook.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class LeaveDetailsAPIView(View):
    def get(self, request, leave_id, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            employee = Employee.objects.get(employee_id=employee_id)
            
            try:
                leave = LeaveApplication.objects.get(leave_id=leave_id, employee=employee)
            except LeaveApplication.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Leave application not found'
                }, status=404)
            
            # Format the leave application data
            leave_data = {
                'leave_id': str(leave.leave_id),
                'leave_type': leave.leave_type,
                'leave_type_display': leave.get_leave_type_display(),
                'start_date': leave.start_date.strftime('%d %b %Y'),
                'end_date': leave.end_date.strftime('%d %b %Y'),
                'reason': leave.reason,
                'status': leave.status,
                'duration': leave.duration,
                'created_at': leave.created_at.strftime('%d %b %Y'),
                'updated_at': leave.updated_at.strftime('%d %b %Y'),
            }
            
            # Add optional fields if they exist
            if leave.document:
                leave_data['document'] = leave.document.url
            
            if leave.reviewed_by:
                leave_data['reviewed_by'] = leave.reviewed_by.name
                leave_data['reviewed_at'] = leave.reviewed_at.strftime('%d %b %Y')
            
            if leave.review_notes:
                leave_data['review_notes'] = leave.review_notes
            
            return JsonResponse({
                'success': True,
                'leave': leave_data
            })
            
        except Exception as e:
            logger.error(f"Error retrieving leave details: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class ReimbursementDetailsAPIView(View):
    def get(self, request, reimbursement_id, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            employee = Employee.objects.get(employee_id=employee_id)
            
            try:
                reimbursement = ReimbursementRequest.objects.get(reimbursement_id=reimbursement_id, employee=employee)
            except ReimbursementRequest.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Reimbursement request not found'
                }, status=404)
            
            # Format the reimbursement request data
            reimbursement_data = {
                'reimbursement_id': str(reimbursement.reimbursement_id),
                'category': reimbursement.category,
                'category_display': reimbursement.get_category_display(),
                'amount': float(reimbursement.amount),
                'currency': reimbursement.currency,
                'expense_date': reimbursement.expense_date.strftime('%d %b %Y'),
                'description': reimbursement.description,
                'status': reimbursement.status,
                'created_at': reimbursement.created_at.strftime('%d %b %Y'),
                'updated_at': reimbursement.updated_at.strftime('%d %b %Y'),
            }
            
            # Add optional fields if they exist
            if reimbursement.receipt:
                reimbursement_data['receipt'] = reimbursement.receipt.url
            
            if reimbursement.approved_by:
                reimbursement_data['approved_by'] = reimbursement.approved_by.name
                reimbursement_data['approved_at'] = reimbursement.approved_at.strftime('%d %b %Y')
            
            if reimbursement.approval_notes:
                reimbursement_data['approval_notes'] = reimbursement.approval_notes
            
            if reimbursement.payment_date:
                reimbursement_data['payment_date'] = reimbursement.payment_date.strftime('%d %b %Y')
            
            if reimbursement.payment_reference:
                reimbursement_data['payment_reference'] = reimbursement.payment_reference
            
            return JsonResponse({
                'success': True,
                'reimbursement': reimbursement_data
            })
            
        except Exception as e:
            logger.error(f"Error retrieving reimbursement details: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class SalarySlipDetailsAPIView(View):
    def get(self, request, salary_slip_id, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            employee = Employee.objects.get(employee_id=employee_id)
            
            try:
                salary_slip = SalarySlip.objects.get(salary_slip_id=salary_slip_id, employee=employee)
            except SalarySlip.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Salary slip not found'
                }, status=404)
            
            # Format the salary slip data
            salary_slip_data = {
                'salary_slip_id': str(salary_slip.salary_slip_id),
                'month': salary_slip.month,
                'year': salary_slip.year,
                'period_display': salary_slip.period_display,
                'basic_salary': float(salary_slip.basic_salary),
                'net_salary': float(salary_slip.net_salary),
                'payment_date': salary_slip.payment_date.strftime('%d %b %Y'),
                'payment_method': salary_slip.payment_method,
                'is_paid': salary_slip.is_paid,
                'allowances': salary_slip.allowances,
                'deductions': salary_slip.deductions,
                'created_at': salary_slip.created_at.strftime('%d %b %Y'),
                'employee_name': employee.name,
                'employee_id': employee.employee_id,
                'department': employee.department.name if hasattr(employee, 'department') and employee.department else None,
                'designation': employee.designation.name if hasattr(employee, 'designation') and employee.designation else None,
            }
            
            # Add optional fields if they exist
            if salary_slip.pdf_file:
                salary_slip_data['pdf_file'] = salary_slip.pdf_file.url
            
            if salary_slip.payment_reference:
                salary_slip_data['payment_reference'] = salary_slip.payment_reference
            
            if salary_slip.notes:
                salary_slip_data['notes'] = salary_slip.notes
            
            return JsonResponse({
                'success': True,
                'salary_slip': salary_slip_data
            })
            
        except Exception as e:
            logger.error(f"Error retrieving salary slip details: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class ReimbursementCancelAPIView(View):
    def post(self, request, reimbursement_id, *args, **kwargs):
        employee_id = request.session.get('employee_id')
        if not employee_id:
            return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
        
        try:
            employee = Employee.objects.get(employee_id=employee_id)
            
            try:
                reimbursement = ReimbursementRequest.objects.get(reimbursement_id=reimbursement_id, employee=employee)
            except ReimbursementRequest.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Reimbursement request not found'
                }, status=404)
            
            # Check if reimbursement can be cancelled
            if reimbursement.status != 'pending':
                return JsonResponse({
                    'success': False,
                    'message': f'Cannot cancel reimbursement request in {reimbursement.status} status'
                }, status=400)
            
            # Update reimbursement status to cancelled (or delete it)
            reimbursement.status = 'cancelled'  # You may need to add 'cancelled' to the status choices
            reimbursement.updated_at = timezone.now()
            reimbursement.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Reimbursement request cancelled successfully'
            })
            
        except Exception as e:
            logger.error(f"Error cancelling reimbursement request: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

class SalaryView(TemplateView):
    template_name = 'hr_management/salary.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = None
        user_session_id = self.request.session.get('user_session_id')

        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None

        if not user:
            context['employees'] = []
            return context

        try:
            companies = Company.objects.filter(user=user)
            company_ids = [c.company_id for c in companies]
            employees = Employee.objects.filter(company__company_id__in=company_ids).select_related('company')
            context['employees'] = employees
        except Exception as e:
            logger.error(f"Error fetching employees for salary view: {e}", exc_info=True)
            context['employees'] = []

        return context

class ExportIncentiveSheetView(View):
    def get(self, request, *args, **kwargs):
        user = None
        user_session_id = request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None

        if not user:
            return HttpResponse("Unauthorized", status=401)

        try:
            companies = Company.objects.filter(user=user)
            company_ids = [c.company_id for c in companies]
            employees = Employee.objects.filter(company_id__in=company_ids)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Incentive Sheet"

            headers = [
                "Employee ID", "Employee Name", "Employee Mobile Number",
                "Incentive Amount", "Incentive Reason"
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = sheet[f"{col_letter}1"]
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)

            for row_num, employee in enumerate(employees, 2):
                sheet[f"A{row_num}"] = str(employee.employee_id)
                sheet[f"B{row_num}"] = employee.employee_name
                sheet[f"C{row_num}"] = employee.phone_number if employee.phone_number else ""
                sheet[f"D{row_num}"] = ""  # Incentive Amount (leave blank)
                sheet[f"E{row_num}"] = ""  # Incentive Reason (leave blank)
            
            for i in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(i)].width = 20

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="incentive_sheet.xlsx"'
            workbook.save(response)
            return response

        except Exception as e:
            logger.error(f"Error exporting incentive sheet: {e}", exc_info=True)
            return HttpResponse("An error occurred during export.", status=500)

class ExportDeductionSheetView(View):
    def get(self, request, *args, **kwargs):
        user = None
        user_session_id = request.session.get('user_session_id')
        if user_session_id:
            try:
                user_session = UserSession.objects.get(id=user_session_id)
                user = user_session.user
            except UserSession.DoesNotExist:
                user = None

        if not user:
            return HttpResponse("Unauthorized", status=401)

        try:
            companies = Company.objects.filter(user=user)
            company_ids = [c.company_id for c in companies]
            employees = Employee.objects.filter(company_id__in=company_ids)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Deductions Sheet"

            headers = [
                "Employee ID", "Employee Name", "Employee Mobile Number",
                "Deduction Amount", "Deduction Reason"
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = sheet[f"{col_letter}1"]
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)

            for row_num, employee in enumerate(employees, 2):
                sheet[f"A{row_num}"] = str(employee.employee_id)
                sheet[f"B{row_num}"] = employee.employee_name
                sheet[f"C{row_num}"] = employee.phone_number if employee.phone_number else ""
                sheet[f"D{row_num}"] = ""  # Deduction Amount (leave blank)
                sheet[f"E{row_num}"] = ""  # Deduction Reason (leave blank)
            
            for i in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(i)].width = 20

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="deductions_sheet.xlsx"'
            workbook.save(response)
            return response

        except Exception as e:
            logger.error(f"Error exporting deductions sheet: {e}", exc_info=True)
            return HttpResponse("An error occurred during export.", status=500)

@method_decorator(csrf_exempt, name='dispatch')
class SendOfferAcceptanceOTPView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        invitation = get_object_or_404(OnboardingInvitation, invitation_id=invitation_id)

        if invitation.status != 'sent':
            return JsonResponse({'success': False, 'error': 'This offer is no longer active.'}, status=400)

        otp = str(random.randint(100000, 999999))
        otp_expiry = timezone.now() + timedelta(minutes=10)

        request.session['offer_acceptance_otp'] = otp
        request.session['offer_acceptance_otp_expiry'] = otp_expiry.isoformat()
        request.session['offer_invitation_id'] = str(invitation.invitation_id)

        try:
            subject = 'Your One-Time Password (OTP) for Offer Acceptance'
            message = f"""
Dear {invitation.name},

To securely accept your job offer for the position of {invitation.designation.name} at our company, please use the following One-Time Password (OTP).

Your OTP is: {otp}

This OTP is valid for 10 minutes.

If you did not request this OTP, please disregard this email.

Best regards,
The HR Team
"""
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [invitation.email],
                fail_silently=False,
            )
            return JsonResponse({'success': True, 'message': 'OTP sent successfully.'})
        except Exception as e:
            logger.error(f"Error sending OTP email for invitation {invitation_id}: {e}")
            return JsonResponse({'success': False, 'error': 'Could not send OTP. Please try again.'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class VerifyOfferAcceptanceOTPView(View):
    def post(self, request, invitation_id, *args, **kwargs):
        try:
            data = json.loads(request.body)
            user_otp = data.get('otp')
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

        stored_otp = request.session.get('offer_acceptance_otp')
        otp_expiry_str = request.session.get('offer_acceptance_otp_expiry')
        session_invitation_id = request.session.get('offer_invitation_id')

        if not all([stored_otp, otp_expiry_str, session_invitation_id]) or session_invitation_id != str(invitation_id):
            return JsonResponse({'success': False, 'error': 'OTP not found or session expired. Please try again.'}, status=400)

        otp_expiry = datetime.fromisoformat(otp_expiry_str)

        if timezone.now() > otp_expiry:
            # Clear expired OTP from session
            if 'offer_acceptance_otp' in request.session:
                del request.session['offer_acceptance_otp']
            if 'offer_acceptance_otp_expiry' in request.session:
                del request.session['offer_acceptance_otp_expiry']
            if 'offer_invitation_id' in request.session:
                del request.session['offer_invitation_id']
            return JsonResponse({'success': False, 'error': 'OTP has expired. Please request a new one.'}, status=400)

        if user_otp == stored_otp:
            # OTP is correct, proceed with accepting the offer
            invitation = get_object_or_404(OnboardingInvitation, invitation_id=invitation_id)
            invitation.status = 'accepted'
            invitation.response_date = timezone.now()
            invitation.save()

            # Clean up session
            if 'offer_acceptance_otp' in request.session:
                del request.session['offer_acceptance_otp']
            if 'offer_acceptance_otp_expiry' in request.session:
                del request.session['offer_acceptance_otp_expiry']
            if 'offer_invitation_id' in request.session:
                del request.session['offer_invitation_id']
            
            form_url = invitation.form_link
            
            return JsonResponse({'success': True, 'message': 'Offer accepted successfully.', 'form_url': form_url})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid OTP. Please try again.'}, status=400)


class ConfigureView(TemplateView):
    template_name = 'hr_management/configure.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class HRConfigurationView(View):
    template_name = 'hr_management/hr_configuration.html'
    
    def get_model(self, config_type):
        if config_type == 'leaves':
            return LeaveType
        elif config_type == 'holidays':
            return Holiday
        elif config_type == 'payouts':
            return PayoutType
        elif config_type == 'deduction_rules':
            return AttendanceRule
        return None

    def get(self, request, config_type):
        model = self.get_model(config_type)
        if not model:
            return HttpResponse("Invalid configuration type", status=404)
        
        user = get_user_from_session(request)
        objects = model.objects.filter(user=user) if user else model.objects.all()
        
        companies = Company.objects.filter(user=user) if user else Company.objects.all()

        context = {
            'config_type': config_type,
            'objects': objects,
            'page_title': f"Manage {config_type.capitalize()}",
            'companies': companies,
        }

        if config_type == 'deduction_rules':
            context['weekdays'] = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        if config_type == 'leaves':
            leave_groups = LeaveGroup.objects.filter(user=user) if user else LeaveGroup.objects.all()
            context['leave_groups'] = leave_groups
            # The 'objects' variable already contains the leave types

        return render(request, self.template_name, context)

    def post(self, request, config_type):
        model = self.get_model(config_type)
        if not model:
            return HttpResponse("Invalid configuration type", status=404)

        user = get_user_from_session(request)
        object_id = request.POST.get('object_id')

        if config_type != 'deduction_rules':
            name = request.POST.get('name')
            if not name:
                return redirect('hr:hr_configuration', config_type=config_type)

        if object_id:
            # Update existing object
            obj = get_object_or_404(model, pk=object_id)
            if user:
                if hasattr(obj, 'user') and obj.user != user:
                    return HttpResponse("Unauthorized", status=403)
            
            if config_type == 'holidays':
                obj.name = request.POST.get('name')
                date = request.POST.get('date')
                if date:
                    obj.date = date
            elif config_type == 'payouts':
                obj.name = request.POST.get('name')
            elif config_type == 'deduction_rules':
                obj.punch_in_time = request.POST.get('punch_in_time')
                obj.punch_out_time = request.POST.get('punch_out_time')
                obj.working_days = ','.join(request.POST.getlist('working_days'))
                company_id = request.POST.get('company')
                if company_id:
                    obj.company_id = company_id
            obj.save()
        else:
            # Create new object
            data = {}
            if config_type == 'holidays':
                data['name'] = request.POST.get('name')
                date = request.POST.get('date')
                if date:
                    data['date'] = date
            elif config_type == 'payouts':
                data['name'] = request.POST.get('name')
            elif config_type == 'deduction_rules':
                data['punch_in_time'] = request.POST.get('punch_in_time')
                data['punch_out_time'] = request.POST.get('punch_out_time')
                data['working_days'] = ','.join(request.POST.getlist('working_days'))
                company_id = request.POST.get('company')
                if company_id:
                    data['company_id'] = company_id
            
            if user:
                data['user'] = user

            
            model.objects.create(**data)

        return redirect('hr:hr_configuration', config_type=config_type)


class HRConfigurationDeleteView(View):
    def get_model(self, config_type):
        if config_type == 'leaves':
            return LeaveType
        elif config_type == 'holidays':
            return Holiday
        elif config_type == 'payouts':
            return PayoutType
        elif config_type == 'deduction_rules':
            return AttendanceRule
        return None

    def post(self, request, config_type, pk):
        model = self.get_model(config_type)
        if not model:
            return HttpResponse("Invalid configuration type", status=404)
        
        user = get_user_from_session(request)
        obj = get_object_or_404(model, pk=pk)
        if user:
            if obj.user != user:
                return HttpResponse("Unauthorized", status=403)
        
        obj.delete()
        return redirect('hr:hr_configuration', config_type=config_type)


class LeaveGroupCreateView(View):
    def post(self, request, *args, **kwargs):
        user = get_user_from_session(request)
        name = request.POST.get('name')
        frequency = request.POST.get('frequency')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        group = LeaveGroup.objects.create(name=name, frequency=frequency, user=user)
        return JsonResponse({'success': True, 'group': {'id': group.group_id, 'name': group.name, 'frequency': group.frequency}})


class LeaveGroupUpdateView(View):
    def post(self, request, group_id, *args, **kwargs):
        user = get_user_from_session(request)
        name = request.POST.get('name')
        frequency = request.POST.get('frequency')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        group = get_object_or_404(LeaveGroup, group_id=group_id, user=user)
        group.name = name
        group.frequency = frequency
        group.save()
        return JsonResponse({'success': True, 'group': {'id': group.group_id, 'name': group.name, 'frequency': group.frequency}})


class LeaveGroupDeleteView(View):
    def post(self, request, group_id, *args, **kwargs):
        user = get_user_from_session(request)
        group = get_object_or_404(LeaveGroup, group_id=group_id, user=user)
        group.delete()
        return JsonResponse({'success': True})


class LeaveGroupRuleCreateView(View):
    def post(self, request, group_id, *args, **kwargs):
        user = get_user_from_session(request)
        leave_type_id = request.POST.get('leave_type_id')
        days = request.POST.get('days')

        if not all([leave_type_id, days]):
            return JsonResponse({'success': False, 'error': 'Leave type and days are required'})

        group = get_object_or_404(LeaveGroup, group_id=group_id, user=user)
        leave_type = get_object_or_404(LeaveType, leave_type_id=leave_type_id, user=user)

        rule = LeaveGroupRule.objects.create(leave_group=group, leave_type=leave_type, days=days)
        return JsonResponse({'success': True, 'rule': {'id': rule.leave_group_rule_id, 'leave_type_name': rule.leave_type.name, 'days': rule.days}})


class LeaveGroupRuleUpdateView(View):
    def post(self, request, rule_id, *args, **kwargs):
        user = get_user_from_session(request)
        days = request.POST.get('days')

        if not days:
            return JsonResponse({'success': False, 'error': 'Days are required'})

        rule = get_object_or_404(LeaveGroupRule, leave_group_rule_id=rule_id, leave_group__user=user)
        rule.days = days
        rule.save()
        return JsonResponse({'success': True, 'rule': {'id': rule.leave_group_rule_id, 'leave_type_name': rule.leave_type.name, 'days': rule.days}})


class LeaveGroupRuleDeleteView(View):
    def post(self, request, rule_id, *args, **kwargs):
        user = get_user_from_session(request)
        rule = get_object_or_404(LeaveGroupRule, leave_group_rule_id=rule_id, leave_group__user=user)
        rule.delete()
        return JsonResponse({'success': True})

@method_decorator(csrf_exempt, name='dispatch')
class ConfigureAttendanceRuleView(View):
    def post(self, request, rule_id):
        try:
            with transaction.atomic():
                rule = AttendanceRule.objects.get(id=rule_id)
                data = json.loads(request.body)
                
                # Clear existing sub-rules
                rule.late_punch_in_rules.all().delete()
                rule.early_punch_out_rules.all().delete()

                # Create new late punch-in rules
                for sub_rule_data in data.get('late_punch_in_rules', []):
                    if not sub_rule_data.get('start_time') or not sub_rule_data.get('rule_option'):
                        continue
                    LatePunchInRule.objects.create(
                        attendance_rule=rule,
                        start_time=sub_rule_data.get('start_time'),
                        end_time=sub_rule_data.get('end_time') or None,
                        rule_option=sub_rule_data.get('rule_option'),
                        deduction_amount=sub_rule_data.get('deduction_amount') if sub_rule_data.get('rule_option') == 'deduct_amount' else None
                    )

                # Create new early punch-out rules
                for sub_rule_data in data.get('early_punch_out_rules', []):
                    if not sub_rule_data.get('time') or not sub_rule_data.get('rule_option'):
                        continue
                    EarlyPunchOutRule.objects.create(
                        attendance_rule=rule,
                        time=sub_rule_data.get('time'),
                        rule_option=sub_rule_data.get('rule_option'),
                        deduction_amount=sub_rule_data.get('deduction_amount') if sub_rule_data.get('rule_option') == 'deduct_amount' else None
                    )

            return JsonResponse({'success': True, 'message': 'Configuration saved successfully.'})
        except AttendanceRule.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Attendance rule not found.'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON format.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

class LeaveGroupCreateView(View):
    def post(self, request, *args, **kwargs):
        user = get_user_from_session(request)
        name = request.POST.get('name')
        frequency = request.POST.get('frequency')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        group = LeaveGroup.objects.create(name=name, frequency=frequency, user=user)
        return JsonResponse({'success': True, 'group': {'id': group.group_id, 'name': group.name, 'frequency': group.frequency}})